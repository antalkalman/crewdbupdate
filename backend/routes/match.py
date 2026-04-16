from __future__ import annotations

from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException

from backend.services.match_service import MatchService
from backend.utils.paths import MASTER_DIR, GCMID_MAP_FILE, CREW_REGISTRY_FILE

router = APIRouter()


def get_match_service() -> MatchService:
    return MatchService()


@router.post("/match/run")
def run_match(service: MatchService = Depends(get_match_service)) -> dict:
    return service.run()


@router.get("/match/status")
def match_status(service: MatchService = Depends(get_match_service)) -> dict:
    return service.status()


@router.post("/match/new_run")
def new_run_match(service: MatchService = Depends(get_match_service)) -> dict:
    result = service.run_new()

    try:
        from datetime import datetime
        from backend.routes.workflow import update_status
        meta = result.get("meta", {})
        update_status("last_match", {
            "timestamp": datetime.now().isoformat(),
            "confirmed": meta.get("confirmed_count", 0),
            "possible": meta.get("possible_count", 0),
            "new_names": meta.get("new_names_count", 0),
            "missing": meta.get("missing_count", 0),
        })
    except Exception:
        pass

    return result


class ConfirmEntry(BaseModel):
    cm_job: str
    cm_id: int


class ConfirmToHelperRequest(BaseModel):
    entries: list[ConfirmEntry]


@router.post("/match/confirm_to_helper")
def confirm_to_helper(request: ConfirmToHelperRequest) -> dict:
    import openpyxl
    import re

    helper_path = MASTER_DIR / "Helper.xlsx"
    if not helper_path.exists():
        raise HTTPException(status_code=404, detail=f"Helper.xlsx not found: {helper_path}")

    try:
        wb = openpyxl.load_workbook(helper_path)
    except PermissionError:
        raise HTTPException(status_code=423, detail="Helper.xlsx is open. Please close it and try again.")

    if "GCMID" not in wb.sheetnames:
        raise HTTPException(status_code=500, detail="Sheet 'GCMID' not found in Helper.xlsx")

    ws = wb["GCMID"]

    if "GCM" not in ws.tables:
        raise HTTPException(status_code=500, detail="Table 'GCM' not found in GCMID sheet")

    # Read formula from first data row (row 2) to reuse exactly
    formula_str = ws.cell(2, 3).value or '=_xlfn.TEXTAFTER(GCM[[#This Row],[CM-Job]],"--")'

    # Collect existing CM-Job values to detect duplicates (skip header row 1)
    existing_cm_jobs: set[str] = set()
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        val = row[0].value
        if val is not None:
            existing_cm_jobs.add(str(val).strip())

    added = 0
    skipped = 0

    for entry in request.entries:
        cm_job = str(entry.cm_job).strip()
        if cm_job in existing_cm_jobs:
            skipped += 1
            continue
        next_row = ws.max_row + 1
        ws.cell(next_row, 1).value = cm_job
        ws.cell(next_row, 2).value = entry.cm_id
        ws.cell(next_row, 3).value = formula_str
        existing_cm_jobs.add(cm_job)
        added += 1

    # Extend the table ref to include newly appended rows
    if added > 0:
        table = ws.tables["GCM"]
        match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", table.ref)
        if match:
            col_start, row_start, col_end, row_end = match.groups()
            new_last_row = int(row_end) + added
            table.ref = f"{col_start}{row_start}:{col_end}{new_last_row}"

    try:
        wb.save(helper_path)
    except PermissionError:
        raise HTTPException(status_code=423, detail="Could not save Helper.xlsx. Please close it and try again.")

    return {"added": added, "skipped": skipped}


class NewNameEntry(BaseModel):
    source_key: str
    name_on_crew_list: str
    project_job_title: str
    actual_title_override: str = ""
    phone: str = ""
    email: str = ""
    general_department: str = ""
    status: str = "Active"
    note: str = ""


class AddNewNamesRequest(BaseModel):
    entries: list[NewNameEntry]


@router.post("/match/add_new_names")
def add_new_names(request: AddNewNamesRequest) -> dict:
    import openpyxl
    import re

    names_path = MASTER_DIR / "Names.xlsx"
    if not names_path.exists():
        raise HTTPException(status_code=404, detail=f"Names.xlsx not found: {names_path}")

    try:
        wb = openpyxl.load_workbook(names_path)
    except PermissionError:
        raise HTTPException(status_code=423, detail="Names.xlsx is open. Please close it and try again.")

    ws = wb.active

    if "Táblázat1" not in ws.tables:
        raise HTTPException(status_code=500, detail="Table 'Táblázat1' not found in Names.xlsx")

    # Find current max CM ID from col A (skip header row 1)
    max_id = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        val = row[0].value
        if val is not None:
            try:
                max_id = max(max_id, int(val))
            except (ValueError, TypeError):
                pass

    added = 0
    new_ids: list[int] = []

    for entry in request.entries:
        new_id = max_id + 1
        max_id = new_id

        # Split name: "Firstname Lastname" → first=everything before last space, sure=last token
        name = (entry.name_on_crew_list or "").strip()
        parts = name.rsplit(" ", 1)
        first_name = parts[0] if len(parts) > 1 else name
        sure_name  = parts[1] if len(parts) > 1 else ""

        # Clean phone: strip non-digits → int or None
        phone_digits = re.sub(r"\D", "", str(entry.phone or ""))
        if not phone_digits or str(entry.phone).strip().lower() == "nan":
            phone_val = None
        else:
            try:
                phone_val = int(phone_digits)
            except ValueError:
                phone_val = None

        # Clean email
        email_val = entry.email.strip() if entry.email and entry.email.strip().lower() != "nan" else None

        # Append row: CM ID, Sure Name, First Name, Nick Name, Actual Title,
        #             Actual Phone, Actual Email, Note, Actual Name (left blank — Excel fills formula)
        next_row = ws.max_row + 1
        ws.cell(next_row, 1).value = new_id
        ws.cell(next_row, 2).value = sure_name
        ws.cell(next_row, 3).value = first_name
        ws.cell(next_row, 4).value = ""
        ws.cell(next_row, 5).value = entry.project_job_title or ""
        ws.cell(next_row, 6).value = phone_val
        ws.cell(next_row, 7).value = email_val
        ws.cell(next_row, 8).value = entry.general_department or ""
        # col I (Actual Name) left blank — Excel applies table formula on open

        new_ids.append(new_id)
        added += 1

    # Extend table ref
    if added > 0:
        table = ws.tables["Táblázat1"]
        match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", table.ref)
        if match:
            col_start, row_start, col_end, row_end = match.groups()
            table.ref = f"{col_start}{row_start}:{col_end}{int(row_end) + added}"

    try:
        wb.save(names_path)
    except PermissionError:
        raise HTTPException(status_code=423, detail="Could not save Names.xlsx. Please close it and try again.")

    return {"added": added, "new_ids": new_ids}


@router.post("/match/confirm_to_gcmid_map")
def confirm_to_gcmid_map(request: ConfirmToHelperRequest) -> dict:
    import openpyxl

    if not GCMID_MAP_FILE.exists():
        raise HTTPException(status_code=404, detail=f"GCMID_Map.xlsx not found: {GCMID_MAP_FILE}")
    try:
        wb = openpyxl.load_workbook(GCMID_MAP_FILE)
    except PermissionError:
        raise HTTPException(status_code=423, detail="GCMID_Map.xlsx is open. Please close it and try again.")

    ws = wb["GCMID_Map"]

    existing_cm_jobs: set[str] = set()
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        val = row[0].value
        if val is not None:
            existing_cm_jobs.add(str(val).strip())

    added = 0
    skipped = 0

    for entry in request.entries:
        cm_job = str(entry.cm_job).strip()
        if cm_job in existing_cm_jobs:
            skipped += 1
            continue
        project = cm_job.split("--", 1)[1] if "--" in cm_job else ""
        next_row = ws.max_row + 1
        ws.cell(next_row, 1).value = cm_job
        ws.cell(next_row, 2).value = entry.cm_id
        ws.cell(next_row, 3).value = project
        ws.cell(next_row, 4).value = "Confirmed"
        existing_cm_jobs.add(cm_job)
        added += 1

    try:
        wb.save(GCMID_MAP_FILE)
    except PermissionError:
        raise HTTPException(status_code=423, detail="Could not save GCMID_Map.xlsx. Please close it and try again.")

    return {"added": added, "skipped": skipped}


@router.post("/match/add_to_registry")
def add_to_registry(request: AddNewNamesRequest) -> dict:
    import openpyxl
    import re

    if not CREW_REGISTRY_FILE.exists():
        raise HTTPException(status_code=404, detail=f"CrewRegistry.xlsx not found: {CREW_REGISTRY_FILE}")
    try:
        wb = openpyxl.load_workbook(CREW_REGISTRY_FILE)
    except PermissionError:
        raise HTTPException(status_code=423, detail="CrewRegistry.xlsx is open. Please close it and try again.")

    ws = wb.active

    # Load TitleMap for General Title lookup
    from backend.utils.paths import TITLEMAP_FILE
    title_conv: dict[str, str] = {}
    try:
        wb_tm = openpyxl.load_workbook(TITLEMAP_FILE, read_only=True)
        ws_tm = wb_tm["Title conv"]
        for row in ws_tm.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                title_conv[str(row[0]).strip()] = str(row[1]).strip()
        wb_tm.close()
    except Exception:
        pass  # If TitleMap unavailable, proceed without lookup

    max_id = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        val = row[0].value
        if val is not None:
            try:
                max_id = max(max_id, int(val))
            except (ValueError, TypeError):
                pass

    added = 0
    new_ids: list[int] = []

    for entry in request.entries:
        new_id = max_id + 1
        max_id = new_id

        name = (entry.name_on_crew_list or "").strip()
        parts = name.rsplit(" ", 1)
        first_name = parts[0] if len(parts) > 1 else name
        sure_name = parts[1] if len(parts) > 1 else ""

        phone_digits = re.sub(r"\D", "", str(entry.phone or ""))
        phone_val = int(phone_digits) if phone_digits and str(entry.phone).strip().lower() != "nan" else None

        BLOCKED_EMAILS = {"pioneer@crewcall.hu"}
        email_val = entry.email.strip() if entry.email and entry.email.strip().lower() != "nan" else None
        if email_val and email_val.lower() in BLOCKED_EMAILS:
            email_val = None

        actual_name = f"{first_name} {sure_name}".strip()

        # Derive General Title: use modal override if provided, else TitleMap lookup
        if entry.actual_title_override:
            general_title = entry.actual_title_override
        else:
            project = entry.source_key.split("--", 1)[1] if "--" in entry.source_key else ""
            title_key = f"{entry.project_job_title}--{project}" if entry.project_job_title and project else ""
            general_title = title_conv.get(title_key, "")

        # Note: only write what the user explicitly provided in the modal
        note_val = entry.note or ""

        next_row = ws.max_row + 1
        ws.cell(next_row, 1).value = new_id
        ws.cell(next_row, 2).value = sure_name
        ws.cell(next_row, 3).value = first_name
        ws.cell(next_row, 4).value = None            # Nick Name
        ws.cell(next_row, 5).value = general_title   # Actual Title
        ws.cell(next_row, 6).value = entry.status    # Status
        ws.cell(next_row, 7).value = phone_val       # Actual Phone
        ws.cell(next_row, 8).value = email_val       # Actual Email
        ws.cell(next_row, 9).value = note_val        # Note
        ws.cell(next_row, 10).value = None           # Last General Title
        ws.cell(next_row, 11).value = None           # Last Department
        ws.cell(next_row, 12).value = None           # Title Flag
        ws.cell(next_row, 13).value = email_val      # Last Email (same as Actual for new records)
        ws.cell(next_row, 14).value = phone_val      # Last Phone (same as Actual for new records)
        ws.cell(next_row, 15).value = None           # Shows Worked
        ws.cell(next_row, 16).value = actual_name    # Actual Name

        new_ids.append(new_id)
        added += 1

    if added > 0:
        table = ws.tables["CrewRegistry"]
        match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", table.ref)
        if match:
            col_start, row_start, col_end, row_end = match.groups()
            table.ref = f"{col_start}{row_start}:{col_end}{int(row_end) + added}"

    # Re-apply inline Status dropdown validation (openpyxl corrupts cross-sheet refs)
    from x_new_combine import ensure_status_validation
    ensure_status_validation(ws)

    try:
        wb.save(CREW_REGISTRY_FILE)
    except PermissionError:
        raise HTTPException(status_code=423, detail="Could not save CrewRegistry.xlsx. Please close it and try again.")

    return {"added": added, "new_ids": new_ids}
