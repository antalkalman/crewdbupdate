from __future__ import annotations

import openpyxl
from pydantic import BaseModel
from fastapi import APIRouter, HTTPException

from backend.utils.paths import CREW_REGISTRY_FILE

router = APIRouter()

EDITABLE_COLS = {"Sure Name", "First Name", "Nick Name", "Actual Title", "Status",
                 "Actual Phone", "Actual Email", "Note"}


@router.get("/registry")
def get_registry() -> dict:
    if not CREW_REGISTRY_FILE.exists():
        raise HTTPException(status_code=404, detail="CrewRegistry.xlsx not found")
    try:
        wb = openpyxl.load_workbook(CREW_REGISTRY_FILE, read_only=True, data_only=True)
    except PermissionError:
        raise HTTPException(status_code=423, detail="CrewRegistry.xlsx is open. Please close it.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {"columns": [], "rows": []}

    header = [str(c) for c in rows[0]]
    records = []
    for row in rows[1:]:
        rec = {}
        for i, col in enumerate(header):
            val = row[i] if i < len(row) else None
            if val is None:
                rec[col] = ""
            elif isinstance(val, bool):
                rec[col] = str(val)
            else:
                rec[col] = val
        records.append(rec)

    return {"columns": header, "rows": records}


class RegistryUpdate(BaseModel):
    cm_id: int
    field: str
    value: str


class RegistrySaveRequest(BaseModel):
    changes: list[RegistryUpdate]


@router.post("/registry/save")
def save_registry(request: RegistrySaveRequest) -> dict:
    if not CREW_REGISTRY_FILE.exists():
        raise HTTPException(status_code=404, detail="CrewRegistry.xlsx not found")
    if not request.changes:
        return {"saved": 0}

    for change in request.changes:
        if change.field not in EDITABLE_COLS:
            raise HTTPException(status_code=400, detail=f"Column '{change.field}' is not editable")

    try:
        wb = openpyxl.load_workbook(CREW_REGISTRY_FILE)
    except PermissionError:
        raise HTTPException(status_code=423, detail="CrewRegistry.xlsx is open. Please close it.")

    ws = wb.active
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    cm_id_col = header.index("CM ID") + 1
    cm_id_to_row: dict[int, int] = {}
    for row_num in range(2, ws.max_row + 1):
        val = ws.cell(row_num, cm_id_col).value
        if val is not None:
            try:
                cm_id_to_row[int(val)] = row_num
            except (ValueError, TypeError):
                pass

    saved = 0
    for change in request.changes:
        row_num = cm_id_to_row.get(change.cm_id)
        if row_num is None:
            continue
        if change.field not in header:
            continue
        col_num = header.index(change.field) + 1
        ws.cell(row_num, col_num).value = change.value if change.value != "" else None
        saved += 1

    # Re-apply Status dropdown validation before saving
    from x_new_combine import ensure_status_validation
    ensure_status_validation(ws)

    try:
        wb.save(CREW_REGISTRY_FILE)
    except PermissionError:
        raise HTTPException(status_code=423, detail="Could not save CrewRegistry.xlsx. Please close it.")

    return {"saved": saved}
