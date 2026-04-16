from __future__ import annotations

import io
import json
from datetime import date

import openpyxl
import pandas as pd
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel

from backend.utils.paths import NEW_MASTER_DIR, SF_ARCHIVE_DIR

router = APIRouter()

ISSUES_STATE_FILE = NEW_MASTER_DIR / "sf_issues_state.json"

REQUIRED_ALL = [
    "Project department", "Project job title", "Project unit",
    "Surname", "Firstname", "Mobile number",
    "Crew list name", "Crew email",
    "Start date", "End date", "Deal type",
    "Project overtime", "Project turnaround", "Project working hour",
]
REQUIRED_BD_SKIP = {"Surname", "Firstname", "Mobile number", "Crew list name", "Crew email"}

FIELD_MESSAGES = {
    "Project department": "Missing department",
    "Project job title": "Missing job title",
    "Project unit": "Missing unit",
    "Surname": "Missing name",
    "Firstname": "Missing name",
    "Mobile number": "Missing phone number",
    "Crew list name": "Missing crew list name",
    "Crew email": "Missing email",
    "Start date": "Missing start date",
    "End date": "Missing end date",
    "Deal type": "Missing deal type",
    "Project overtime": "Missing overtime",
    "Project turnaround": "Missing turnaround",
    "Project working hour": "Missing working hours",
}

VISIBLE_COLS = [
    "Sf number", "Crew member id", "State", "Project",
    "Project department", "Project job title", "Project unit",
    "Crew list name", "Surname", "Firstname",
    "Mobile number", "Crew email",
    "Start date", "End date", "Deal type",
    "Project overtime", "Project turnaround", "Project working hour",
    "Daily fee", "Weekly fee",
]

EXPORT_COLS = [
    "Sf number", "Crew member id", "Crew list name",
    "Project job title", "Project department", "Note",
]


def _is_blank(val) -> bool:
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in {"", "nan", "none", "n/a", "-"}


def _has_fee(row: dict) -> bool:
    for col in ("Daily fee", "Weekly fee"):
        val = row.get(col)
        if not _is_blank(val):
            try:
                if float(str(val).replace(",", "")) > 0:
                    return True
            except (ValueError, TypeError):
                pass
    return False


def _find_issues(row: dict, today: date) -> str:
    issues: list[str] = []
    sf_num = str(row.get("Sf number", ""))
    sf_type = sf_num[:2] if len(sf_num) >= 2 else ""
    state = str(row.get("State", "")).lower().strip()

    # Timing check — not signed but start date has passed
    start_raw = row.get("Start date")
    if state not in {"accepted", "signed"} and not _is_blank(start_raw):
        try:
            start_dt = pd.to_datetime(start_raw).date()
            if start_dt < today:
                days = (today - start_dt).days
                issues.append(f"Not signed \u2014 start date was {days} days ago")
        except Exception:
            pass

    # Completeness checks — only for accepted/signed
    if state in {"accepted", "signed"}:
        required = list(REQUIRED_ALL)
        if sf_type == "BD":
            required = [f for f in required if f not in REQUIRED_BD_SKIP]

        seen_messages: set[str] = set()
        for field in required:
            if _is_blank(row.get(field)):
                msg = FIELD_MESSAGES.get(field, f"Missing {field}")
                if msg not in seen_messages:
                    issues.append(msg)
                    seen_messages.add(msg)

        if not _has_fee(row):
            issues.append("No fee set")

    return " \u00b7 ".join(issues)


@router.post("/sf_issues/run")
def run_sf_issues() -> dict:
    sf_files = sorted(SF_ARCHIVE_DIR.glob("SFlist_*.xlsx")) if SF_ARCHIVE_DIR.exists() else []
    if not sf_files:
        raise HTTPException(status_code=404, detail="No SFlist file found in SF_Archive/")
    sf_path = sf_files[-1]

    try:
        df = pd.read_excel(sf_path, dtype=str, engine="openpyxl")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not read SFlist: {exc}") from exc

    all_columns = list(df.columns)

    # Filter draft/paused
    if "State" in df.columns:
        df = df[~df["State"].fillna("").str.lower().isin(["draft", "paused"])]

    today = date.today()
    rows: list[dict] = []
    for _, row in df.iterrows():
        row_dict = row.to_dict()
        note = _find_issues(row_dict, today)
        record = {col: str(row_dict.get(col, "") or "").strip() for col in all_columns}
        record["Note"] = note
        record["has_issues"] = bool(note)
        record["checked"] = False
        rows.append(record)

    # Merge persisted state
    persisted: dict = {}
    if ISSUES_STATE_FILE.exists():
        try:
            persisted = json.loads(ISSUES_STATE_FILE.read_text())
        except Exception:
            pass

    for row in rows:
        key = row.get("Sf number", "")
        if key in persisted:
            row["checked"] = persisted[key].get("checked", False)
            if persisted[key].get("note_edited"):
                row["Note"] = persisted[key]["note"]

    projects = sorted({r.get("Project", "") for r in rows if r.get("Project")})
    issue_count = sum(1 for r in rows if r["has_issues"])

    return {
        "rows": rows,
        "columns": [c for c in VISIBLE_COLS if c in all_columns] + ["Note"],
        "all_columns": all_columns,
        "source_file": sf_path.name,
        "total": len(rows),
        "with_issues": issue_count,
        "projects": projects,
    }


class SaveStateRequest(BaseModel):
    changes: list[dict]


@router.post("/sf_issues/save_state")
def save_sf_issues_state(request: SaveStateRequest) -> dict:
    persisted: dict = {}
    if ISSUES_STATE_FILE.exists():
        try:
            persisted = json.loads(ISSUES_STATE_FILE.read_text())
        except Exception:
            pass
    for change in request.changes:
        sf_num = change.get("sf_number")
        if sf_num:
            persisted[sf_num] = {
                "checked": change.get("checked", False),
                "note": change.get("note", ""),
                "note_edited": change.get("note_edited", False),
            }
    ISSUES_STATE_FILE.write_text(json.dumps(persisted, indent=2, ensure_ascii=False))
    return {"saved": len(request.changes)}


class ExportRequest(BaseModel):
    rows: list[dict]
    filename: str = "SF_Issues"


@router.post("/sf_issues/export")
def export_sf_issues(request: ExportRequest):
    by_project: dict[str, list] = {}
    for row in request.rows:
        proj = row.get("Project", "Unknown")
        by_project.setdefault(proj, []).append(row)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="BDD7EE")

    for proj, proj_rows in sorted(by_project.items()):
        ws = wb.create_sheet(title=proj[:31])
        ws.append(EXPORT_COLS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in proj_rows:
            ws.append([row.get(c, "") for c in EXPORT_COLS])
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 60

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = "".join(c for c in request.filename if c.isalnum() or c in "._- ")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )
