from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException

from backend.dependencies import get_title_service
from backend.services.title_service import TitleService
from backend.utils.paths import BASE_DIR, NEW_MASTER_DIR

router = APIRouter()

CREWINDEX_PATH = NEW_MASTER_DIR / "CrewIndex.xlsx"


class ApplyMappingsRequest(BaseModel):
    rows: list[dict]


@router.get("/unmapped_titles")
def get_unmapped_titles(service: TitleService = Depends(get_title_service)) -> dict:
    return service.get_unmapped_titles()


@router.post("/apply_title_mappings")
def apply_title_mappings(
    payload: ApplyMappingsRequest,
    service: TitleService = Depends(get_title_service),
) -> dict[str, int]:
    return service.append_title_mappings_to_helper(service.helper_path, payload.rows)


@router.get("/general_titles")
def get_general_titles() -> dict:
    titles: list[str] = []
    try:
        wb = openpyxl.load_workbook(
            BASE_DIR / "New_Master_Database" / "TitleMap.xlsx", read_only=True,
        )
        ws = wb["General Title"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                titles.append(str(row[1]).strip())
        wb.close()
    except Exception:
        pass
    return {"titles": sorted(set(titles))}


@router.get("/general_titles_with_dept")
def get_general_titles_with_dept() -> dict:
    items: list[dict] = []
    try:
        wb = openpyxl.load_workbook(
            BASE_DIR / "New_Master_Database" / "TitleMap.xlsx", read_only=True,
        )
        ws = wb["General Title"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            dept = str(row[0]).strip() if row[0] else None
            title = str(row[1]).strip() if row[1] else None
            if dept and title:
                items.append({"dept": dept, "title": title})
        wb.close()
    except Exception:
        pass
    return {"items": items}


@router.get("/title_conflicts")
def get_title_conflicts(service: TitleService = Depends(get_title_service)) -> list[dict]:
    """Compute title conflicts dynamically from CrewIndex.xlsx.

    A conflict = same Title--Project key maps to multiple different General Titles
    across different rows (e.g. from different historical files).
    Only returns keys not yet in TitleMap.
    """
    if not CREWINDEX_PATH.exists():
        return []

    try:
        wb = openpyxl.load_workbook(CREWINDEX_PATH, read_only=True, data_only=True)
    except Exception:
        return []

    ws = wb["CrewIndex"]
    header = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    def find_col(name: str) -> int | None:
        for i, h in enumerate(header):
            if h == name:
                return i
        return None

    col_pjt = find_col("Project job title")
    col_proj = find_col("Project")
    col_gt = find_col("General Title")
    col_dept = find_col("Project department")

    if col_pjt is None or col_proj is None or col_gt is None:
        wb.close()
        return []

    # Collect all GT values per Title--Project key
    pairs: dict[str, dict] = defaultdict(
        lambda: {"gts": set(), "counter": Counter(), "dept": "", "count": 0}
    )

    for row in ws.iter_rows(min_row=2, values_only=True):
        pjt = row[col_pjt] if col_pjt < len(row) else None
        proj = row[col_proj] if col_proj < len(row) else None
        gt = row[col_gt] if col_gt < len(row) else None

        if not pjt or not proj or not gt:
            continue

        pjt_s = str(pjt).strip()
        proj_s = str(proj).strip()
        gt_s = str(gt).strip()
        if not pjt_s or not proj_s or not gt_s:
            continue

        key = f"{pjt_s}--{proj_s}"
        pairs[key]["gts"].add(gt_s)
        pairs[key]["counter"][gt_s] += 1
        pairs[key]["count"] += 1
        if col_dept is not None and col_dept < len(row) and row[col_dept]:
            pairs[key]["dept"] = str(row[col_dept]).strip()

    wb.close()

    # Load existing Title conv keys to identify already-resolved
    existing_keys, _ = service.read_helper_title_conv(service.titlemap_path)

    # Only report keys with multiple GTs (actual conflicts)
    result = []
    for key, info in sorted(pairs.items()):
        if len(info["gts"]) <= 1:
            continue  # Not a conflict

        majority = info["counter"].most_common(1)[0][0]
        candidates = sorted(info["gts"])
        pjt_part = key.split("--", 1)[0] if "--" in key else key
        proj_part = key.split("--", 1)[1] if "--" in key else ""

        result.append({
            "title_project": key,
            "project_job_title": pjt_part,
            "project": proj_part,
            "candidates": candidates,
            "majority_vote": majority,
            "row_count": info["count"],
            "department": info["dept"],
            "already_resolved": key in existing_keys,
        })

    return result
