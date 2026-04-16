from __future__ import annotations

import json

import openpyxl
from fastapi import APIRouter

from backend.utils.paths import NEW_MASTER_DIR, SF_ARCHIVE_DIR

router = APIRouter()
STATUS_FILE = NEW_MASTER_DIR / "status.json"


def update_status(section: str, data: dict) -> None:
    """Persist a workflow section to status.json. Never raises."""
    try:
        status: dict = {}
        if STATUS_FILE.exists():
            status = json.loads(STATUS_FILE.read_text())
        status[section] = data
        STATUS_FILE.write_text(json.dumps(status, indent=2, ensure_ascii=False))
    except Exception as exc:
        print(f"[workflow] WARNING: could not update status.json: {exc}")


@router.get("/workflow/status")
def get_workflow_status() -> dict:
    # Load persisted status
    status: dict = {}
    if STATUS_FILE.exists():
        try:
            status = json.loads(STATUS_FILE.read_text())
        except Exception:
            pass

    # Latest SFlist file
    sf_files = sorted(SF_ARCHIVE_DIR.glob("SFlist_*.xlsx")) if SF_ARCHIVE_DIR.exists() else []
    latest_sf = sf_files[-1] if sf_files else None

    # Unmapped titles count
    unmapped_count = 0
    try:
        from backend.services.title_service import TitleService
        svc = TitleService()
        result = svc.get_unmapped_titles()
        unmapped_count = len(result.get("rows", []))
    except Exception:
        pass

    # Unresolved crew count from GCMID_Map
    unresolved_count = 0
    gcmid_map_path = NEW_MASTER_DIR / "GCMID_Map.xlsx"
    try:
        wb = openpyxl.load_workbook(gcmid_map_path, read_only=True)
        if "Unresolved" in wb.sheetnames:
            ws = wb["Unresolved"]
            unresolved_count = sum(1 for _ in ws.iter_rows(min_row=2))
        wb.close()
    except Exception:
        pass

    # CrewRegistry stats — find columns by header name dynamically
    registry_path = NEW_MASTER_DIR / "CrewRegistry.xlsx"
    registry_stats: dict = {}
    try:
        wb = openpyxl.load_workbook(registry_path, read_only=True)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        idx_status = header.index("Status") if "Status" in header else None
        idx_title = header.index("Actual Title") if "Actual Title" in header else None
        idx_shows = header.index("Shows Worked") if "Shows Worked" in header else None
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        def get_status(r):
            return str(r[idx_status] or "").strip() if idx_status is not None else ""

        registry_stats = {
            "total": len(rows),
            "active": sum(1 for r in rows if get_status(r) == "Active"),
            "retired": sum(1 for r in rows if get_status(r) == "Retired"),
            "foreign": sum(1 for r in rows if get_status(r) == "Foreign"),
            "external": sum(1 for r in rows if get_status(r) == "External"),
            "no_actual_title": sum(1 for r in rows if idx_title is not None and not r[idx_title]),
            "no_shows": sum(1 for r in rows if idx_shows is not None and not r[idx_shows]),
        }
        wb.close()
    except Exception:
        pass

    return {
        "last_export": status.get("last_export"),
        "last_pipeline": status.get("last_pipeline"),
        "last_match": status.get("last_match"),
        "live": {
            "latest_sf_file": latest_sf.name if latest_sf else None,
            "latest_sf_mtime": latest_sf.stat().st_mtime if latest_sf else None,
            "unmapped_titles": unmapped_count,
            "unresolved_crew": unresolved_count,
            "registry": registry_stats,
        },
    }
