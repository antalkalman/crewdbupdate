"""Add Actual Phone + Actual Email columns to CrewRegistry.xlsx.

Inserts cols 7-8 between Status (col 6) and Note (was col 7, now col 9).
Populates from Master_database/Names.xlsx where CM IDs match.
Rebuilds the Excel Table and dropdown validations.
Result: 16 columns instead of 14.
"""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

REGISTRY = Path("New_Master_Database/CrewRegistry.xlsx")
NAMES = Path("Master_database/Names.xlsx")

NEW_HEADERS = [
    "CM ID",              # 1
    "Sure Name",          # 2
    "First Name",         # 3
    "Nick Name",          # 4
    "Actual Title",       # 5
    "Status",             # 6
    "Actual Phone",       # 7  ← NEW
    "Actual Email",       # 8  ← NEW
    "Note",               # 9  (was 7)
    "Last General Title", # 10 (was 8)
    "Last Department",    # 11 (was 9)
    "Title Flag",         # 12 (was 10)
    "Last Email",         # 13 (was 11)
    "Last Phone",         # 14 (was 12)
    "Shows Worked",       # 15 (was 13)
    "Actual Name",        # 16 (was 14)
]

# Old column order (14 cols):
# 1:CM ID  2:Sure Name  3:First Name  4:Nick Name  5:Actual Title  6:Status
# 7:Note   8:Last General Title  9:Last Department  10:Title Flag
# 11:Last Email  12:Last Phone  13:Shows Worked  14:Actual Name

# Mapping: new_col_index (0-based) → old_col_index (0-based)
OLD_TO_NEW = {
    0: 0,   # CM ID
    1: 1,   # Sure Name
    2: 2,   # First Name
    3: 3,   # Nick Name
    4: 4,   # Actual Title
    5: 5,   # Status
    6: None,  # Actual Phone — from Names.xlsx
    7: None,  # Actual Email — from Names.xlsx
    8: 6,   # Note (was 7)
    9: 7,   # Last General Title (was 8)
    10: 8,  # Last Department (was 9)
    11: 9,  # Title Flag (was 10)
    12: 10, # Last Email (was 11)
    13: 11, # Last Phone (was 12)
    14: 12, # Shows Worked (was 13)
    15: 13, # Actual Name (was 14)
}


def load_names_contacts() -> dict:
    """Load CM ID → {actual_phone, actual_email} from Names.xlsx."""
    contacts = {}
    if not NAMES.exists():
        print(f"  WARNING: {NAMES} not found — Actual Phone/Email will be empty")
        return contacts

    wb = openpyxl.load_workbook(NAMES, read_only=True, data_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    def find(name):
        for i, h in enumerate(header):
            if h == name:
                return i
        return None

    col_cmid = find("CM ID")
    col_phone = find("Actual Phone")
    col_email = find("Actual Email")

    for row in ws.iter_rows(min_row=2, values_only=True):
        cm_id_val = row[col_cmid] if col_cmid is not None else None
        if cm_id_val is None:
            continue
        try:
            cm_id = int(cm_id_val)
        except (ValueError, TypeError):
            continue

        phone = None
        if col_phone is not None:
            pv = row[col_phone]
            if pv is not None:
                try:
                    phone = int(re.sub(r"\D", "", str(pv))) if str(pv).strip() else None
                except (ValueError, TypeError):
                    phone = None

        email = None
        if col_email is not None:
            ev = row[col_email]
            if ev and str(ev).strip().lower() not in ("", "none", "nan"):
                email = str(ev).strip().lower()

        contacts[cm_id] = {"actual_phone": phone, "actual_email": email}

    wb.close()
    print(f"  Names.xlsx: {len(contacts)} records loaded")
    return contacts


def main():
    print("migrate_add_actual_contacts.py\n")

    contacts = load_names_contacts()

    # Load existing registry
    wb_old = openpyxl.load_workbook(REGISTRY, data_only=True)
    ws_old = wb_old.active
    old_header = [cell.value for cell in ws_old[1]]
    print(f"  Old schema: {len(old_header)} columns")

    old_rows = []
    for row in ws_old.iter_rows(min_row=2, values_only=True):
        old_rows.append(list(row))
    wb_old.close()

    # Also preserve hidden sheets
    wb_full = openpyxl.load_workbook(REGISTRY)
    title_list_data = []
    if "_TitleList" in wb_full.sheetnames:
        ws_tl = wb_full["_TitleList"]
        for row in ws_tl.iter_rows(values_only=True):
            if row[0]:
                title_list_data.append(str(row[0]).strip())
    wb_full.close()

    # Build new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CrewRegistry"

    # Header
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    for col_idx, name in enumerate(NEW_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font

    # Auto column fill (cols 10-15: Last General Title through Shows Worked)
    auto_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    # Title Flag fill
    flag_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

    phone_filled = 0
    email_filled = 0

    for old_row in old_rows:
        new_row = [None] * 16
        for new_idx, old_idx in OLD_TO_NEW.items():
            if old_idx is not None and old_idx < len(old_row):
                new_row[new_idx] = old_row[old_idx]

        # Get CM ID for Names.xlsx lookup
        cm_id = None
        try:
            cm_id = int(new_row[0]) if new_row[0] is not None else None
        except (ValueError, TypeError):
            pass

        if cm_id and cm_id in contacts:
            c = contacts[cm_id]
            if c["actual_phone"] is not None:
                new_row[6] = c["actual_phone"]
                phone_filled += 1
            if c["actual_email"] is not None:
                new_row[7] = c["actual_email"]
                email_filled += 1

        row_num = ws.max_row + 1
        for col_idx, val in enumerate(new_row, 1):
            ws.cell(row=row_num, column=col_idx, value=val)

        # Apply auto column fill (cols 10-15)
        for col in range(10, 16):
            ws.cell(row=row_num, column=col).fill = auto_fill
        # Title Flag yellow fill (col 12)
        ws.cell(row=row_num, column=12).fill = flag_fill

    ws.freeze_panes = "A2"

    # Column widths
    widths = {
        1: 8, 2: 16, 3: 16, 4: 12, 5: 22, 6: 10, 7: 15, 8: 25,
        9: 20, 10: 22, 11: 20, 12: 12, 13: 25, 14: 15, 15: 40, 16: 22,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Create Excel Table
    last_row = ws.max_row
    table_ref = f"A1:P{last_row}"
    table = Table(displayName="CrewRegistry", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Recreate _TitleList hidden sheet
    if title_list_data:
        ws_tl = wb.create_sheet("_TitleList")
        for i, title in enumerate(title_list_data, 1):
            ws_tl.cell(row=i, column=1, value=title)
        ws_tl.sheet_state = "hidden"

    # Actual Title dropdown (col E) referencing _TitleList
    if title_list_data:
        title_count = len(title_list_data)
        dv_title = DataValidation(
            type="list",
            formula1=f"=_TitleList!$A$1:$A${title_count}",
            allow_blank=True,
        )
        dv_title.add(f"E2:E{last_row}")
        ws.add_data_validation(dv_title)

    # Status dropdown (col F) — inline list
    dv_status = DataValidation(
        type="list",
        formula1='"Active,Retired,Foreign,External"',
        allow_blank=True,
    )
    dv_status.error = "Please select a valid status"
    dv_status.errorTitle = "Invalid Status"
    dv_status.add(f"F2:F{last_row}")
    ws.add_data_validation(dv_status)

    wb.save(REGISTRY)

    print(f"""
CrewRegistry migration — added Actual Phone + Actual Email
  Total records:        {len(old_rows):,}
  Actual Phone filled:  {phone_filled:,}  (from Names.xlsx)
  Actual Email filled:  {email_filled:,}  (from Names.xlsx)
  Output: {REGISTRY} (now {len(NEW_HEADERS)} columns)

New schema:""")
    for i, h in enumerate(NEW_HEADERS, 1):
        marker = " ← NEW" if h in ("Actual Phone", "Actual Email") else ""
        print(f"  {i:2d}: {h}{marker}")


if __name__ == "__main__":
    main()
