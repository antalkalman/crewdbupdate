"""One-time migration: auto-fill clean Title conv mappings from Historical files."""

from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

HELPER = Path("Master_database/Helper.xlsx")
HIST_DIR = Path("New_Master_Database/Historical")
CONFLICTS_OUT = Path("New_Master_Database/title_conv_conflicts.xlsx")

PROTECTED = {
    "Practical Dimmer Operator--BETAMAX",
    "Practical Chargehand--3BP",
    "Leather Maker--Sharp",
    "Playback Operator--Sharp",
    "Moving Unit Assistant--Sharp",
    "Practical Chargehand--3BP Season 3",
}


def is_empty(val):
    return val is None or str(val).strip() in ("", "None", "nan")


def find_col(header, name):
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


def main():
    # ── Step 1: Load reference data ──
    wb_h = openpyxl.load_workbook(HELPER, read_only=True, data_only=True)

    # Existing Title conv keys
    ws_tc = wb_h["Title conv"]
    existing_keys = set()
    for row in ws_tc.iter_rows(min_row=2, values_only=True):
        key = row[0]
        if key and str(key).strip():
            existing_keys.add(str(key).strip())
    initial_count = len(existing_keys)

    # Valid General Titles
    ws_gt = wb_h["General Title"]
    valid_titles = set()
    for row in ws_gt.iter_rows(min_row=2, values_only=True):
        title = row[1]  # Title column (B)
        if title:
            valid_titles.add(str(title).strip())

    wb_h.close()
    print(f"Existing Title conv keys: {initial_count}")
    print(f"Valid General Titles: {len(valid_titles)}")

    # ── Step 2: Collect pairs from Historical files ──
    files = sorted(
        p for p in HIST_DIR.glob("*.xlsx")
        if not p.name.startswith("~$") and p.name != "Historical_data_250601.xlsx"
    )

    # key -> { gt_set, gt_counter, rows list for conflict detail }
    pairs = defaultdict(lambda: {"gts": set(), "counter": Counter(), "dept": "", "count": 0})

    for path in files:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        header = rows[0]
        col_proj = find_col(header, "Project")
        col_pjt = find_col(header, "Project job title")
        col_gt = find_col(header, "General Title")
        col_dept = find_col(header, "Project department")

        if col_pjt is None or col_proj is None or col_gt is None:
            print(f"  Skipping {path.name}: missing required columns")
            continue

        for row in rows[1:]:
            pjt = row[col_pjt]
            proj = row[col_proj]
            gt = row[col_gt]

            if is_empty(pjt) or is_empty(proj) or is_empty(gt):
                continue

            pjt_s = str(pjt).strip()
            proj_s = str(proj).strip()
            gt_s = str(gt).strip()
            key = f"{pjt_s}--{proj_s}"

            pairs[key]["gts"].add(gt_s)
            pairs[key]["counter"][gt_s] += 1
            pairs[key]["count"] += 1
            if col_dept is not None and not is_empty(row[col_dept]):
                pairs[key]["dept"] = str(row[col_dept]).strip()

    print(f"Unique Title-Project keys from Historical: {len(pairs)}")

    # ── Categorise ──
    already_mapped = 0
    protected_count = 0
    invalid_gt_list = []
    conflict_list = []
    clean_pairs = []  # (key, general_title)

    for key, info in sorted(pairs.items()):
        if key in existing_keys:
            already_mapped += 1
            continue
        if key in PROTECTED:
            protected_count += 1
            continue

        gts = info["gts"]
        if len(gts) == 1:
            gt = next(iter(gts))
            if gt in valid_titles:
                clean_pairs.append((key, gt))
            else:
                invalid_gt_list.append((key, gt))
        else:
            # Conflict: multiple GTs
            majority = info["counter"].most_common(1)[0][0]
            candidates = " / ".join(sorted(gts))
            pjt_part = key.split("--", 1)[0] if "--" in key else key
            proj_part = key.split("--", 1)[1] if "--" in key else ""
            conflict_list.append({
                "key": key,
                "pjt": pjt_part,
                "project": proj_part,
                "candidates": candidates,
                "majority": majority,
                "count": info["count"],
                "dept": info["dept"],
            })

    # ── Step 3: Write clean pairs to Helper.xlsx ──
    print(f"\nWriting {len(clean_pairs)} clean pairs to Helper.xlsx Title conv...")
    wb_w = openpyxl.load_workbook(HELPER)
    ws_w = wb_w["Title conv"]

    for key, gt in clean_pairs:
        ws_w.append([key, gt])

    try:
        wb_w.save(HELPER)
        print("  Helper.xlsx saved successfully.")
    except PermissionError:
        print("  ERROR: Could not save Helper.xlsx — please close the file and re-run.")
        return

    # ── Step 4: Write conflicts log ──
    out = openpyxl.Workbook()
    ws_c = out.active
    ws_c.title = "Conflicts"
    ws_c.append(["Title-Project", "Project job title", "Project",
                 "Candidate General Titles", "Majority Vote", "Row Count", "Department"])

    conflict_list.sort(key=lambda r: (r["project"], r["key"]))
    for c in conflict_list:
        ws_c.append([c["key"], c["pjt"], c["project"],
                      c["candidates"], c["majority"], c["count"], c["dept"]])

    CONFLICTS_OUT.parent.mkdir(parents=True, exist_ok=True)
    out.save(CONFLICTS_OUT)

    # ── Summary ──
    final_count = initial_count + len(clean_pairs)
    print(f"\nTitle conv migration")
    print(f"  Already mapped (skipped):  {already_mapped}")
    print(f"  Protected (skipped):       {protected_count}")
    print(f"  Invalid General Title:     {len(invalid_gt_list)}")
    print(f"  Conflicts (logged):        {len(conflict_list)}")
    print(f"  Clean pairs written:       {len(clean_pairs):,}")
    print(f"\n  Helper.xlsx Title conv: was {initial_count:,} → now {final_count:,} entries")
    print(f"\n  Conflicts log: {CONFLICTS_OUT}")

    if invalid_gt_list:
        print(f"\n  Sample invalid GTs:")
        for key, gt in invalid_gt_list[:10]:
            print(f"    {key} → '{gt}'")


if __name__ == "__main__":
    main()
