"""One-time migration: build CrewRegistry.xlsx from Names.xlsx + Historical derived data."""

import glob
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

HELPER = Path("Master_database/Helper.xlsx")
NAMES = Path("Master_database/Names.xlsx")
HIST_DIR = Path("New_Master_Database/Historical")
SF_ARCHIVE = Path("SF_Archive")
GCMID_MAP = Path("New_Master_Database/GCMID_Map.xlsx")
OUTPUT = Path("New_Master_Database/CrewRegistry.xlsx")


def is_empty(val):
    return val is None or str(val).strip() in ("", "None", "nan")


def clean_int(val):
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return int(s)


def digits_only(val) -> str:
    if is_empty(val):
        return ""
    return re.sub(r"\D", "", str(val).strip())


def find_col(header, name):
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


# ── Step 1: Lookup tables from Helper ──

def load_helper():
    wb = openpyxl.load_workbook(HELPER, read_only=True, data_only=True)

    # General Title sheet
    ws_gt = wb["General Title"]
    title_to_dept = {}
    title_to_id = {}
    all_titles = []
    for row in ws_gt.iter_rows(min_row=2, values_only=True):
        dept, title, dept_id, title_id = row[0], row[1], row[2], row[3]
        if title:
            t = str(title).strip()
            title_to_dept[t] = str(dept).strip() if dept else ""
            if title_id is not None:
                title_to_id[t] = int(title_id)
            all_titles.append(t)

    # FProjects sheet
    ws_fp = wb["FProjects"]
    project_dates = {}
    for row in ws_fp.iter_rows(min_row=2, values_only=True):
        proj, start_date = row[0], row[1]
        if proj and start_date is not None:
            project_dates[str(proj).strip()] = int(start_date)

    wb.close()
    return title_to_dept, title_to_id, all_titles, project_dates


# ── Step 2: Read Historical + SFlist data ──

def load_gcmid_map():
    """Load CM-Job → GCMID mapping for SFlist rows."""
    path = GCMID_MAP
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["GCMID_Map"]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cm_job, cm_id = row[0], row[1]
        if cm_job and cm_id is not None:
            mapping[str(cm_job).strip()] = int(cm_id)
    wb.close()
    return mapping


def read_historical_files(project_dates):
    """Read all Historical files and return rows grouped by GCMID."""
    files = sorted(
        p for p in HIST_DIR.glob("*.xlsx")
        if not p.name.startswith("~$") and p.name != "Historical_data_250601.xlsx"
    )

    # gcmid -> list of dicts
    data = defaultdict(list)

    for path in files:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        header = rows[0]
        col_gcmid = find_col(header, "GCMID")
        col_proj = find_col(header, "Project")
        col_gt = find_col(header, "General Title")
        col_email = find_col(header, "Email")
        col_crew_email = find_col(header, "Crew email")
        col_phone = find_col(header, "Mobile number")

        for row in rows[1:]:
            gcmid_val = row[col_gcmid] if col_gcmid is not None else None
            if is_empty(gcmid_val):
                continue
            gcmid = clean_int(gcmid_val)
            project = str(row[col_proj]).strip() if row[col_proj] else ""
            gen_title = str(row[col_gt]).strip() if col_gt is not None and not is_empty(row[col_gt]) else ""

            email = ""
            if col_email is not None and not is_empty(row[col_email]):
                email = str(row[col_email]).strip().lower()
            elif col_crew_email is not None and not is_empty(row[col_crew_email]):
                email = str(row[col_crew_email]).strip().lower()

            phone_digits = digits_only(row[col_phone] if col_phone is not None else None)

            proj_sort = project_dates.get(project, 0)

            data[gcmid].append({
                "project": project,
                "general_title": gen_title,
                "email": email,
                "phone": phone_digits,
                "project_sort": proj_sort,
            })

    return data


def read_sflist(project_dates, gcmid_map):
    """Read latest SFlist and return rows grouped by GCMID (resolved via map)."""
    pattern = str(SF_ARCHIVE / "SFlist_*.xlsx")
    files = sorted(glob.glob(pattern))
    if not files:
        return {}

    latest = files[-1]
    print(f"Reading SFlist: {Path(latest).name}")

    wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    col_cm = find_col(header, "Crew member id")
    col_proj = find_col(header, "Project")
    col_email = find_col(header, "Email")
    col_crew_email = find_col(header, "Crew email")
    col_phone = find_col(header, "Mobile number")

    data = defaultdict(list)

    for row in rows[1:]:
        cm_val = row[col_cm] if col_cm is not None else None
        proj_val = row[col_proj] if col_proj is not None else None
        if is_empty(cm_val) or is_empty(proj_val):
            continue

        cm_str = str(cm_val).strip()
        project = str(proj_val).strip()
        cm_job = f"{cm_str}--{project}"

        gcmid = gcmid_map.get(cm_job)
        if gcmid is None:
            continue

        email = ""
        if col_email is not None and not is_empty(row[col_email]):
            email = str(row[col_email]).strip().lower()
        elif col_crew_email is not None and not is_empty(row[col_crew_email]):
            email = str(row[col_crew_email]).strip().lower()

        phone_digits = digits_only(row[col_phone] if col_phone is not None else None)
        proj_sort = project_dates.get(project, 0)

        data[gcmid].append({
            "project": project,
            "general_title": "",  # SFlist has no General Title
            "email": email,
            "phone": phone_digits,
            "project_sort": proj_sort,
        })

    return data


def derive_fields(grouped, title_to_dept):
    """From grouped rows per GCMID, derive shows_worked, last_general_title, etc."""
    derived = {}

    for gcmid, rows in grouped.items():
        # Sort by project_sort descending
        rows_sorted = sorted(rows, key=lambda r: r["project_sort"], reverse=True)

        # shows_worked: unique projects, most recent first
        seen_projects = []
        seen_set = set()
        for r in rows_sorted:
            if r["project"] and r["project"] not in seen_set:
                seen_projects.append(r["project"])
                seen_set.add(r["project"])
        shows_worked = " / ".join(seen_projects)

        # last_general_title: from highest project_sort, first non-empty
        last_general_title = ""
        for r in rows_sorted:
            if r["general_title"]:
                last_general_title = r["general_title"]
                break

        last_department = title_to_dept.get(last_general_title, "")

        # last_email: from highest project_sort that has email
        last_email = ""
        for r in rows_sorted:
            if r["email"]:
                last_email = r["email"]
                break

        # last_phone: from highest project_sort that has phone
        last_phone = None
        for r in rows_sorted:
            if r["phone"]:
                try:
                    last_phone = int(r["phone"])
                except ValueError:
                    pass
                break

        derived[gcmid] = {
            "shows_worked": shows_worked,
            "last_general_title": last_general_title,
            "last_department": last_department,
            "last_email": last_email,
            "last_phone": last_phone,
        }

    return derived


# ── Step 3: Read Names.xlsx ──

def load_names():
    wb = openpyxl.load_workbook(NAMES, read_only=True, data_only=True)
    ws = wb.active
    header = list(next(ws.iter_rows(max_row=1, values_only=True)))

    col_cmid = find_col(header, "CM ID")
    col_sure = find_col(header, "Sure Name")
    col_first = find_col(header, "First Name")
    col_nick = find_col(header, "Nick Name")
    col_title = find_col(header, "Actual Title")
    col_phone = find_col(header, "Actual Phone")
    col_email = find_col(header, "Actual Email")
    col_note = find_col(header, "Note")

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cm_id_val = row[col_cmid]
        if is_empty(cm_id_val):
            continue
        try:
            cm_id = clean_int(cm_id_val)
        except (ValueError, TypeError):
            continue

        sure = row[col_sure] if col_sure is not None and not is_empty(row[col_sure]) else ""
        first = row[col_first] if col_first is not None and not is_empty(row[col_first]) else ""
        nick = row[col_nick] if col_nick is not None and not is_empty(row[col_nick]) else None
        actual_title = row[col_title] if col_title is not None and not is_empty(row[col_title]) else ""
        actual_phone = row[col_phone] if col_phone is not None and not is_empty(row[col_phone]) else None
        actual_email = row[col_email] if col_email is not None and not is_empty(row[col_email]) else ""
        note = row[col_note] if col_note is not None and not is_empty(row[col_note]) else ""

        # Actual Name
        actual_name = f"{first} {sure}".strip()
        if nick:
            actual_name += f" '{nick}'"

        records.append({
            "cm_id": cm_id,
            "sure_name": sure,
            "first_name": first,
            "nick_name": nick,
            "actual_title": actual_title,
            "actual_phone": actual_phone,
            "actual_email": actual_email,
            "note": note,
            "actual_name": actual_name,
        })

    wb.close()
    return records


# ── Step 4: Merge & write ──

def main():
    print("Loading Helper lookups...")
    title_to_dept, title_to_id, all_titles, project_dates = load_helper()
    print(f"  Titles: {len(all_titles)}, Projects with dates: {len(project_dates)}")

    print("Loading GCMID map for SFlist resolution...")
    gcmid_map = load_gcmid_map()
    print(f"  GCMID map entries: {len(gcmid_map)}")

    print("Reading Historical files...")
    hist_data = read_historical_files(project_dates)
    print(f"  GCMIDs from Historical: {len(hist_data)}")

    print("Reading latest SFlist...")
    sf_data = read_sflist(project_dates, gcmid_map)
    print(f"  GCMIDs from SFlist: {len(sf_data)}")

    # Merge SF into Historical data
    for gcmid, rows in sf_data.items():
        hist_data[gcmid].extend(rows)

    print("Deriving fields...")
    derived = derive_fields(hist_data, title_to_dept)
    print(f"  Derived records: {len(derived)}")

    print("Loading Names.xlsx...")
    names_records = load_names()
    print(f"  Names records: {len(names_records)}")

    # Sort by CM ID
    names_records.sort(key=lambda r: r["cm_id"])

    # Stats
    stats = {
        "total": len(names_records),
        "actual_title_filled": 0,
        "actual_title_blank": 0,
        "last_gt_filled": 0,
        "last_dept_filled": 0,
        "last_email_filled": 0,
        "last_phone_filled": 0,
        "shows_filled": 0,
        "phonebook_only": 0,
    }

    # Build output rows
    output_rows = []
    for rec in names_records:
        cm_id = rec["cm_id"]
        d = derived.get(cm_id, {})

        last_gt = d.get("last_general_title", "")
        last_dept = d.get("last_department", "")
        last_email = d.get("last_email", "")
        last_phone = d.get("last_phone", None)
        shows = d.get("shows_worked", "")

        if rec["actual_title"]:
            stats["actual_title_filled"] += 1
        else:
            stats["actual_title_blank"] += 1
        if last_gt:
            stats["last_gt_filled"] += 1
        if last_dept:
            stats["last_dept_filled"] += 1
        if last_email:
            stats["last_email_filled"] += 1
        if last_phone:
            stats["last_phone_filled"] += 1
        if shows:
            stats["shows_filled"] += 1
        else:
            stats["phonebook_only"] += 1

        output_rows.append([
            cm_id,                          # A: CM ID
            rec["sure_name"],               # B: Sure Name
            rec["first_name"],              # C: First Name
            rec["nick_name"],               # D: Nick Name
            rec["actual_title"] or None,    # E: Actual Title
            False,                          # F: Retired
            rec["note"] or None,            # G: Note
            last_gt or None,                # H: Last General Title
            last_dept or None,              # I: Last Department
            None,                           # J: Title Flag
            last_email or None,             # K: Last Email
            last_phone,                     # L: Last Phone
            shows or None,                  # M: Shows Worked
            rec["actual_name"],             # N: Actual Name
        ])

    # Write output
    print("Writing CrewRegistry.xlsx...")
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "CrewRegistry"

    headers = [
        "CM ID", "Sure Name", "First Name", "Nick Name", "Actual Title",
        "Retired", "Note", "Last General Title", "Last Department",
        "Title Flag", "Last Email", "Last Phone", "Shows Worked", "Actual Name",
    ]
    ws.append(headers)

    for row in output_rows:
        ws.append(row)

    # ── Formatting ──

    # Header style
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    for col in range(1, 15):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto columns H-M (8-13): light grey fill
    grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    last_row = len(output_rows) + 1
    for row_idx in range(2, last_row + 1):
        for col_idx in [8, 9, 11, 12, 13]:  # H, I, K, L, M
            ws.cell(row=row_idx, column=col_idx).fill = grey_fill
        ws.cell(row=row_idx, column=10).fill = yellow_fill  # J: Title Flag

    # Column widths
    widths = {
        1: 8, 2: 20, 3: 20, 4: 20, 5: 25, 6: 12, 7: 30,
        8: 25, 9: 25, 10: 12, 11: 30, 12: 15, 13: 60, 14: 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Excel Table
    table_ref = f"A1:N{last_row}"
    table = Table(displayName="CrewRegistry", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    # Hidden sheet for title validation list
    ws_titles = out.create_sheet("_TitleList")
    for i, t in enumerate(all_titles, start=1):
        ws_titles.cell(row=i, column=1, value=t)
    ws_titles.sheet_state = "hidden"

    # Data validation for Actual Title (col E)
    title_count = len(all_titles)
    dv_title = DataValidation(
        type="list",
        formula1=f"=_TitleList!$A$1:$A${title_count}",
        allow_blank=True,
    )
    dv_title.error = "Please select a valid title"
    dv_title.errorTitle = "Invalid Title"
    ws.add_data_validation(dv_title)
    dv_title.add(f"E2:E{last_row}")

    # Data validation for Retired (col F)
    dv_retired = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_retired)
    dv_retired.add(f"F2:F{last_row}")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    out.save(OUTPUT)

    # Print summary
    print(f"\nCrewRegistry.xlsx")
    print(f"  Total records:              {stats['total']:,}")
    print(f"  Actual Title filled:        {stats['actual_title_filled']:,}")
    print(f"  Actual Title blank:         {stats['actual_title_blank']:,}")
    print(f"  Last General Title filled:  {stats['last_gt_filled']:,}")
    print(f"  Last Department filled:     {stats['last_dept_filled']:,}")
    print(f"  Last Email filled:          {stats['last_email_filled']:,}")
    print(f"  Last Phone filled:          {stats['last_phone_filled']:,}")
    print(f"  Shows Worked filled:        {stats['shows_filled']:,}")
    print(f"  Phone book only (no shows): {stats['phonebook_only']:,}")
    print(f"  Retired (all False):        {stats['total']:,}")
    print(f"\nOutput: {OUTPUT}")


if __name__ == "__main__":
    main()
