"""One-time migration: build unified GCMID map from Historical files + Helper GCMID sheet."""

from pathlib import Path
import openpyxl

HELPER = Path("Master_database/Helper.xlsx")
HIST_DIR = Path("New_Master_Database/Historical")
OUTPUT = Path("New_Master_Database/GCMID_Map.xlsx")


def clean_gcmid(val):
    """Convert GCMID to int, stripping whitespace and .0 suffix."""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return int(s)


def is_empty(val):
    return val is None or str(val).strip() in ("", "None", "nan")


def find_col(header, name):
    for i, h in enumerate(header):
        if h == name:
            return i
    raise ValueError(f"Column '{name}' not found in header: {header[:10]}...")


def main():
    # ── Source 1: Helper GCMID sheet ──
    wb_h = openpyxl.load_workbook(HELPER, read_only=True, data_only=True)
    ws_h = wb_h["GCMID"]
    helper_map = {}  # CM-Job -> int GCMID
    helper_count = 0
    for row in ws_h.iter_rows(min_row=2, values_only=True):
        cm_job = row[0]
        cm_id = row[1]
        if is_empty(cm_job) or is_empty(cm_id):
            continue
        cm_job = str(cm_job).strip()
        helper_map[cm_job] = clean_gcmid(cm_id)
        helper_count += 1
    wb_h.close()

    # ── Source 2: Historical files ──
    hist_files = sorted(
        p for p in HIST_DIR.glob("*.xlsx")
        if not p.name.startswith("~$") and p.name != "Historical_data_250601.xlsx"
    )

    hist_map = {}       # CM-Job -> (gcmid, project, source_file)
    unresolved = []     # rows with no GCMID
    conflicts = []      # same CM-Job, different GCMID
    skipped = []        # rows with no CM ID
    total_hist_rows = 0

    for path in hist_files:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        header = rows[0]
        col_cm = find_col(header, "Crew member id")
        col_proj = find_col(header, "Project")
        col_gcmid = find_col(header, "GCMID")
        col_crewlist = find_col(header, "Crew list name")

        data = rows[1:]
        total_hist_rows += len(data)

        for row in data:
            cm_id_val = row[col_cm]
            project = row[col_proj]
            gcmid_val = row[col_gcmid]
            crewlist = row[col_crewlist]

            # Skip if no CM ID
            if is_empty(cm_id_val):
                skipped.append((project, crewlist, path.name, "No CM ID"))
                continue

            cm_id_str = str(cm_id_val).strip()
            proj_str = str(project).strip()
            cm_job = f"{cm_id_str}--{proj_str}"

            # Skip if no GCMID
            if is_empty(gcmid_val):
                unresolved.append((cm_job, proj_str, crewlist, path.name, "No GCMID in source"))
                continue

            gcmid_int = clean_gcmid(gcmid_val)

            if cm_job in hist_map:
                existing_gcmid = hist_map[cm_job][0]
                if existing_gcmid != gcmid_int:
                    conflicts.append((cm_job, existing_gcmid, gcmid_int, crewlist, path.name))
                # same or conflict: keep first seen
            else:
                hist_map[cm_job] = (gcmid_int, proj_str, path.name)

    # ── Merge: Historical first, then Helper overlay ──
    final_map = {}  # CM-Job -> (gcmid, project, source)

    # Start with Historical
    for cm_job, (gcmid, proj, src) in hist_map.items():
        final_map[cm_job] = (gcmid, proj, "Historical")

    # Track overlap stats
    from_helper_only = 0
    from_hist_only = 0
    in_both = 0

    for cm_job, gcmid in helper_map.items():
        proj = cm_job.split("--", 1)[1] if "--" in cm_job else ""
        if cm_job in final_map:
            in_both += 1
            final_map[cm_job] = (gcmid, proj, "Helper")  # Helper wins
        else:
            from_helper_only += 1
            final_map[cm_job] = (gcmid, proj, "Helper")

    from_hist_only = len(hist_map) - in_both

    # ── Write output ──
    out = openpyxl.Workbook()

    # Sheet 1: GCMID_Map
    ws1 = out.active
    ws1.title = "GCMID_Map"
    ws1.append(["CM-Job", "CM ID", "Project", "Source"])
    for cm_job in sorted(final_map, key=lambda k: (final_map[k][1], k)):
        gcmid, proj, source = final_map[cm_job]
        ws1.append([cm_job, gcmid, proj, source])

    # Sheet 2: Unresolved
    ws2 = out.create_sheet("Unresolved")
    ws2.append(["CM-Job", "Project", "Crew list name", "Source file", "Note"])
    for row in unresolved:
        ws2.append(list(row))

    # Sheet 3: Conflicts
    ws3 = out.create_sheet("Conflicts")
    ws3.append(["CM-Job", "GCMID kept", "GCMID discarded", "Crew list name", "Source file"])
    for row in conflicts:
        ws3.append(list(row))

    # Sheet 4: Skipped
    ws4 = out.create_sheet("Skipped")
    ws4.append(["Project", "Crew list name", "Source file", "Note"])
    for row in skipped:
        ws4.append(list(row))

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    out.save(OUTPUT)

    # ── Summary ──
    print("Sources processed:")
    print(f"  Helper GCMID sheet:     {helper_count:,} entries")
    print(f"  Historical files:       {len(hist_files)} files, {total_hist_rows:,} rows")
    print()
    print(f"Unified map:              {len(final_map):,} entries")
    print(f"  From Helper only:       {from_helper_only:,}")
    print(f"  From Historical only:   {from_hist_only:,}")
    print(f"  In both (Helper wins):  {in_both:,}")
    print()
    print("Issues logged:")
    print(f"  Unresolved (no GCMID):  {len(unresolved)}")
    print(f"  Conflicts:              {len(conflicts)}")
    print(f"  Skipped (no CM ID):     {len(skipped)}")
    print()
    print(f"Output: {OUTPUT}")


if __name__ == "__main__":
    main()
