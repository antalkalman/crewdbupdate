"""One-time migration: copy Title conv + General Title sheets to TitleMap.xlsx."""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font

HELPER = Path("Master_database/Helper.xlsx")
OUTPUT = Path("New_Master_Database/TitleMap.xlsx")

SHEETS = {
    "Title conv": {"widths": {1: 50, 2: 30}},
    "General Title": {"widths": {1: 25, 2: 35, 3: 15, 4: 15}},
}


def main():
    wb_src = openpyxl.load_workbook(HELPER, read_only=True, data_only=True)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name, cfg in SHEETS.items():
        ws_src = wb_src[sheet_name]
        ws_out = wb_out.create_sheet(title=sheet_name)

        row_count = 0
        for row in ws_src.iter_rows(values_only=True):
            ws_out.append(list(row))
            row_count += 1

        # Bold header + freeze
        for cell in ws_out[1]:
            cell.font = Font(bold=True)
        ws_out.freeze_panes = "A2"

        # Column widths
        from openpyxl.utils import get_column_letter
        for col, w in cfg["widths"].items():
            ws_out.column_dimensions[get_column_letter(col)].width = w

        data_rows = row_count - 1
        print(f"  {sheet_name}: {data_rows:,} entries")

    wb_src.close()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(OUTPUT)

    print(f"\nOutput: {OUTPUT}")


if __name__ == "__main__":
    print("TitleMap.xlsx")
    main()
