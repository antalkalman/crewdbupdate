"""One-time fix: in Historical files, replace GCMID-style CM IDs with GCMID + actual GCMID column value."""

from pathlib import Path
import openpyxl

HIST_DIR = Path("New_Master_Database/Historical")


def is_empty(val):
    return val is None or str(val).strip() in ("", "None", "nan")


def main():
    files = sorted(
        p for p in HIST_DIR.glob("*.xlsx")
        if not p.name.startswith("~$") and p.name != "Historical_data_250601.xlsx"
    )

    total_replaced = 0
    total_skipped_no_gcmid = 0

    for path in files:
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # Find columns by header name
        header = [cell.value for cell in ws[1]]
        col_cm = header.index("Crew member id") + 1      # 1-based for openpyxl
        col_gcmid = header.index("GCMID") + 1

        replaced = 0
        skipped = 0

        for row_idx in range(2, ws.max_row + 1):
            cm_val = ws.cell(row=row_idx, column=col_cm).value
            if cm_val is None:
                continue
            cm_str = str(cm_val).strip()
            if not cm_str.startswith("GCMID"):
                continue

            gcmid_val = ws.cell(row=row_idx, column=col_gcmid).value
            if is_empty(gcmid_val):
                skipped += 1
                continue

            # Clean GCMID: strip .0
            gcmid_str = str(gcmid_val).strip()
            if gcmid_str.endswith(".0"):
                gcmid_str = gcmid_str[:-2]

            new_cm = f"GCMID{gcmid_str}"
            ws.cell(row=row_idx, column=col_cm).value = new_cm
            replaced += 1

        wb.save(path)
        total_replaced += replaced
        total_skipped_no_gcmid += skipped

        if replaced or skipped:
            print(f"{path.name}: {replaced} replaced, {skipped} skipped (no GCMID)")

    print(f"\nTotal: {total_replaced} CM IDs replaced, {total_skipped_no_gcmid} skipped (no GCMID)")


if __name__ == "__main__":
    main()
