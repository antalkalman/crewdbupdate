"""One-time fix: rewrite Historical split files replacing formulas with cached values."""

from pathlib import Path
import openpyxl

DIR = Path("New_Master_Database/Historical")


def main():
    files = sorted(
        p for p in DIR.glob("*.xlsx")
        if not p.name.startswith("~$") and p.name != "Historical_data_250601.xlsx"
    )
    total_rows = 0
    total_formulas = 0

    for path in files:
        wb = openpyxl.load_workbook(path, data_only=True)
        out = openpyxl.Workbook()
        # Remove the default sheet created by Workbook()
        out.remove(out.active)

        file_rows = 0
        file_formulas = 0

        for ws in wb.worksheets:
            ws_out = out.create_sheet(title=ws.title)
            for row in ws.iter_rows(values_only=True):
                for val in row:
                    if isinstance(val, str) and val.startswith("="):
                        file_formulas += 1
                        print(f"  WARNING unresolved formula in {path.name}: {val[:80]}")
                ws_out.append(list(row))
                file_rows += 1

        wb.close()
        out.save(path)

        data_rows = file_rows - 1  # minus header
        total_rows += data_rows
        total_formulas += file_formulas
        print(f"Fixed: {path.name} — {data_rows} rows, {ws_out.max_column} cols")

    print(f"\nFixed {len(files)} files")
    print(f"Total rows written: {total_rows:,}")
    print(f"Any unresolved formulas: {total_formulas}")


if __name__ == "__main__":
    main()
