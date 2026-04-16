"""Fix CrewRegistry.xlsx Status validation.

Replaces the _StatusList sheet-reference validation with an inline list
validation that openpyxl preserves correctly on save.
Removes the _StatusList hidden sheet.
"""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

REGISTRY = Path("New_Master_Database/CrewRegistry.xlsx")
STATUS_LIST = '"Active,Retired,Foreign,External"'


def main():
    wb = openpyxl.load_workbook(REGISTRY)
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    col_idx = header.index("Status") + 1
    col_letter = openpyxl.utils.get_column_letter(col_idx)

    # Remove ALL existing data validations on the Status column
    to_remove = []
    for dv in ws.data_validations.dataValidation:
        for rng in dv.sqref.ranges:
            if str(rng).startswith(col_letter):
                to_remove.append(dv)
                break
    for dv in to_remove:
        ws.data_validations.dataValidation.remove(dv)

    # Add inline list validation
    dv = DataValidation(
        type="list",
        formula1=STATUS_LIST,
        allow_blank=True,
    )
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
    dv.add(cell_range)
    ws.add_data_validation(dv)

    # Remove the _StatusList hidden sheet if it exists
    if "_StatusList" in wb.sheetnames:
        del wb["_StatusList"]
        print("Removed _StatusList hidden sheet")

    wb.save(REGISTRY)
    print(f"Inline validation applied to {cell_range}")
    print(f"Formula: {STATUS_LIST}")


if __name__ == "__main__":
    main()
