"""One-time migration: split Historical_data_250601.xlsx into one file per project."""

import re
from collections import defaultdict
from pathlib import Path

import openpyxl

SRC = Path("New_Master_Database/Historical/Historical_data_250601.xlsx")
DST = Path("New_Master_Database/Historical")

def sanitise(name: str) -> str:
    s = str(name).replace(" ", "_")
    return re.sub(r"[^A-Za-z0-9_\-]", "", s)

def main():
    wb = openpyxl.load_workbook(SRC, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    data = rows[1:]

    groups: dict[str, list] = defaultdict(list)
    for row in data:
        project = row[3]  # "Project" column
        groups[project].append(row)

    DST.mkdir(parents=True, exist_ok=True)

    total = 0
    for project in sorted(groups):
        fname = f"Historical_{sanitise(project)}.xlsx"
        path = DST / fname
        out = openpyxl.Workbook()
        ws_out = out.active
        ws_out.append(list(header))
        for row in groups[project]:
            ws_out.append(list(row))
        out.save(path)
        count = len(groups[project])
        total += count
        print(f"{fname}: {count} rows")

    print(f"\n{len(groups)} projects, {total} total data rows")

if __name__ == "__main__":
    main()
