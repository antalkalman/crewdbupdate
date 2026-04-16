"""Migrate CrewRegistry.xlsx: Retired (boolean) → Status (text dropdown).

Values: Active, Retired, Foreign, External
Adds hidden _StatusList sheet for Excel dropdown validation.
"""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

REGISTRY = Path("New_Master_Database/CrewRegistry.xlsx")
STATUS_VALUES = ["Active", "Retired", "Foreign", "External"]


def main():
    wb = openpyxl.load_workbook(REGISTRY)
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    # Find Retired column dynamically
    try:
        col_idx = header.index("Retired") + 1  # 1-based
    except ValueError:
        print("ERROR: 'Retired' column not found in header:", header)
        return

    # Rename header
    ws.cell(row=1, column=col_idx).value = "Status"

    # Migrate values
    counts = {"Active": 0, "Retired": 0}
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val is True or str(val).strip().upper() == "TRUE":
            cell.value = "Retired"
            counts["Retired"] += 1
        else:
            cell.value = "Active"
            counts["Active"] += 1

    total = counts["Active"] + counts["Retired"]

    # Remove existing data validation on this column (if any)
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    to_remove = []
    for dv in ws.data_validations.dataValidation:
        for rng in dv.sqref.ranges:
            if str(rng).startswith(col_letter):
                to_remove.append(dv)
                break
    for dv in to_remove:
        ws.data_validations.dataValidation.remove(dv)

    # Create hidden _StatusList sheet
    if "_StatusList" in wb.sheetnames:
        del wb["_StatusList"]
    ws_list = wb.create_sheet("_StatusList")
    for i, val in enumerate(STATUS_VALUES, start=1):
        ws_list.cell(row=i, column=1).value = val
    ws_list.sheet_state = "hidden"

    # Add dropdown validation referencing the hidden sheet
    dv = DataValidation(
        type="list",
        formula1="=_StatusList!$A$1:$A$4",
        allow_blank=True,
    )
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    dv.prompt = "Select status"
    dv.promptTitle = "Status"
    cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
    dv.add(cell_range)
    ws.add_data_validation(dv)

    wb.save(REGISTRY)
    print(f"""Status migration complete:
  Active:  {counts['Active']}
  Retired: {counts['Retired']}
  Total:   {total}

Hidden sheet _StatusList created with: {STATUS_VALUES}
Dropdown validation applied to {cell_range}""")


if __name__ == "__main__":
    main()
