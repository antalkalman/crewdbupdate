"""One-time verification: score every GCMID assignment in Historical files against Names_plus."""

import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import Levenshtein
import openpyxl

NAMES_PLUS = Path("Master_database/Names_plus.xlsx")
HELPER = Path("Master_database/Helper.xlsx")
HIST_DIR = Path("New_Master_Database/Historical")
OUTPUT = Path("New_Master_Database/GCMID_Verification.xlsx")


# ── Utilities ──


def is_empty(val):
    return val is None or str(val).strip() in ("", "None", "nan")


def clean_int(val):
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return int(s)


def strip_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn")


def tokenise(name: str) -> list[str]:
    if not name:
        return []
    return strip_accents(name).lower().split()


def digits_only(val) -> str:
    if is_empty(val):
        return ""
    return re.sub(r"\D", "", str(val).strip())


def find_col(header, name):
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


def best_token_score(tokens_a: list[str], tokens_b: list[str]) -> float:
    """Best cumulative match score across token pairs."""
    if not tokens_a or not tokens_b:
        return 0.0
    total = 0.0
    used_b = set()
    for ta in tokens_a:
        best = 0.0
        best_idx = -1
        for j, tb in enumerate(tokens_b):
            if j in used_b:
                continue
            d = Levenshtein.distance(ta, tb)
            if d == 0:
                s = 1.0
            elif d == 1:
                s = 0.5
            elif d == 2:
                s = 0.25
            else:
                s = 0.0
            if s > best:
                best = s
                best_idx = j
        if best > 0 and best_idx >= 0:
            used_b.add(best_idx)
            total += best
    return total


def email_score(hist_email: str, names_emails: list[str]) -> float:
    if not hist_email:
        return 0.0
    h = hist_email.lower().strip()
    best = 0.0
    for ne in names_emails:
        if not ne:
            continue
        n = ne.lower().strip()
        d = Levenshtein.distance(h, n)
        if d == 0:
            return 1.0
        elif d == 1:
            best = max(best, 0.5)
    return best


def phone_score(hist_phone: str, names_phones: list[str]) -> float:
    if not hist_phone:
        return 0.0
    best = 0.0
    for np in names_phones:
        if not np:
            continue
        d = Levenshtein.distance(hist_phone, np)
        if d == 0:
            return 1.0
        elif d == 1:
            best = max(best, 0.5)
    return best


def confidence_label(score: float, in_names: bool) -> str:
    if not in_names:
        return "Check"
    if score >= 2.5:
        return "Strong"
    if score >= 1.5:
        return "Good"
    if score >= 1.0:
        return "Weak"
    return "Poor"


# ── Main ──


def main():
    # Step 1: Load Names_plus
    wb_n = openpyxl.load_workbook(NAMES_PLUS, read_only=True, data_only=True)
    ws_n = wb_n["NamesMasterTable"]
    n_header = list(next(ws_n.iter_rows(max_row=1, values_only=True)))
    nc = {name: find_col(n_header, name) for name in [
        "CM ID", "Sure Name", "First Name", "Nick Name", "Actual Name",
        "Actual Title", "Last General Title", "Actual Phone", "Last Phone Number",
        "Actual Email", "Last Email", "Shows Worked",
    ]}

    names = {}
    for row in ws_n.iter_rows(min_row=2, values_only=True):
        cm_id_val = row[nc["CM ID"]]
        if is_empty(cm_id_val):
            continue
        cm_id = clean_int(cm_id_val)
        sure = row[nc["Sure Name"]] if nc["Sure Name"] is not None else None
        first = row[nc["First Name"]] if nc["First Name"] is not None else None
        nick = row[nc["Nick Name"]] if nc["Nick Name"] is not None else None
        actual_name = row[nc["Actual Name"]] if nc["Actual Name"] is not None else None
        # Build name tokens from actual_name + nick
        name_tokens = tokenise(str(actual_name) if actual_name else "")
        if nick and not is_empty(nick):
            name_tokens += tokenise(str(nick))

        names[cm_id] = {
            "actual_name": actual_name,
            "name_tokens": name_tokens,
            "last_general_title": row[nc["Last General Title"]] if nc["Last General Title"] is not None else None,
            "actual_phone": digits_only(row[nc["Actual Phone"]]),
            "last_phone": digits_only(row[nc["Last Phone Number"]]),
            "actual_email": row[nc["Actual Email"]] if nc["Actual Email"] is not None else None,
            "last_email": row[nc["Last Email"]] if nc["Last Email"] is not None else None,
        }
    wb_n.close()
    print(f"Names_plus loaded: {len(names)} records")

    # Load title→department map from Helper
    wb_h = openpyxl.load_workbook(HELPER, read_only=True, data_only=True)
    ws_gt = wb_h["General Title"]
    title_to_dept = {}
    for row in ws_gt.iter_rows(min_row=2, values_only=True):
        dept, title = row[0], row[1]
        if title:
            title_to_dept[str(title).strip().lower()] = str(dept).strip().lower() if dept else ""
    wb_h.close()
    print(f"Title→Department map: {len(title_to_dept)} entries")

    # Step 2: Load all Historical files
    hist_files = sorted(
        p for p in HIST_DIR.glob("*.xlsx")
        if not p.name.startswith("~$") and p.name != "Historical_data_250601.xlsx"
    )

    results = []
    total_scored = 0

    for path in hist_files:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        header = rows[0]
        col_cm = find_col(header, "Crew member id")
        col_proj = find_col(header, "Project")
        col_gcmid = find_col(header, "GCMID")
        col_crewlist = find_col(header, "Crew list name")
        col_email = find_col(header, "Email")
        col_crew_email = find_col(header, "Crew email")
        col_phone = find_col(header, "Mobile number")
        col_proj_dept = find_col(header, "Project department")
        col_gen_title = find_col(header, "General Title")

        for row in rows[1:]:
            cm_val = row[col_cm]
            gcmid_val = row[col_gcmid] if col_gcmid is not None else None

            if is_empty(cm_val) or is_empty(gcmid_val):
                continue

            cm_str = str(cm_val).strip()
            proj = str(row[col_proj]).strip() if row[col_proj] else ""
            cm_job = f"{cm_str}--{proj}"
            gcmid = clean_int(gcmid_val)
            crewlist = str(row[col_crewlist]).strip() if row[col_crewlist] else ""

            # Email: try Email first, fall back to Crew email
            hist_email = ""
            if col_email is not None and not is_empty(row[col_email]):
                hist_email = str(row[col_email]).strip()
            elif col_crew_email is not None and not is_empty(row[col_crew_email]):
                hist_email = str(row[col_crew_email]).strip()

            hist_phone = digits_only(row[col_phone] if col_phone is not None else None)

            # Department: from General Title via lookup, fall back to Project department
            hist_dept = ""
            if col_gen_title is not None and not is_empty(row[col_gen_title]):
                gt = str(row[col_gen_title]).strip().lower()
                hist_dept = title_to_dept.get(gt, "")
            if not hist_dept and col_proj_dept is not None and not is_empty(row[col_proj_dept]):
                hist_dept = str(row[col_proj_dept]).strip().lower()

            # Step 3: Score
            in_names = gcmid in names
            if not in_names:
                results.append({
                    "cm_job": cm_job, "project": proj, "source": path.name,
                    "crewlist": crewlist, "gcmid": gcmid,
                    "names_record": "", "names_title": "",
                    "name_sc": 0.0, "email_sc": 0.0, "phone_sc": 0.0, "dept_sc": 0.0,
                    "final": 0.0, "confidence": "Check",
                    "note": "GCMID not in Names",
                })
                total_scored += 1
                continue

            rec = names[gcmid]
            # Name score
            crew_tokens = tokenise(crewlist)
            raw_name = best_token_score(crew_tokens, rec["name_tokens"])
            name_sc = min(raw_name * 1.5, 1.5)

            # Email score
            em_sc = email_score(hist_email, [rec["actual_email"], rec["last_email"]])

            # Phone score
            ph_sc = phone_score(hist_phone, [rec["actual_phone"], rec["last_phone"]])

            # Dept score
            d_sc = 0.0
            if hist_dept and rec["last_general_title"]:
                names_dept = title_to_dept.get(str(rec["last_general_title"]).strip().lower(), "")
                if hist_dept and names_dept and hist_dept == names_dept:
                    d_sc = 0.5

            final = name_sc + em_sc + ph_sc + d_sc
            conf = confidence_label(final, True)

            results.append({
                "cm_job": cm_job, "project": proj, "source": path.name,
                "crewlist": crewlist, "gcmid": gcmid,
                "names_record": rec["actual_name"] or "",
                "names_title": rec["last_general_title"] or "",
                "name_sc": round(name_sc, 2), "email_sc": round(em_sc, 2),
                "phone_sc": round(ph_sc, 2), "dept_sc": round(d_sc, 2),
                "final": round(final, 2), "confidence": conf,
                "note": "",
            })
            total_scored += 1

    # Sort: Poor/Check first, then ascending score
    conf_order = {"Check": 0, "Poor": 1, "Weak": 2, "Good": 3, "Strong": 4}
    results.sort(key=lambda r: (conf_order.get(r["confidence"], 5), r["final"]))

    # Step 4: Write output
    out = openpyxl.Workbook()

    # Sheet 1: All Results
    ws1 = out.active
    ws1.title = "All Results"
    hdr = ["CM-Job", "Project", "Source File", "Crew List Name", "GCMID",
           "Names Record", "Names Title", "Name Score", "Email Score",
           "Phone Score", "Dept Score", "Final Score", "Confidence", "Note"]
    ws1.append(hdr)
    for r in results:
        ws1.append([r["cm_job"], r["project"], r["source"], r["crewlist"],
                     r["gcmid"], r["names_record"], r["names_title"],
                     r["name_sc"], r["email_sc"], r["phone_sc"], r["dept_sc"],
                     r["final"], r["confidence"], r["note"]])

    # Sheet 2: Needs Review
    ws2 = out.create_sheet("Needs Review")
    ws2.append(hdr)
    for r in results:
        if r["confidence"] in ("Poor", "Check"):
            ws2.append([r["cm_job"], r["project"], r["source"], r["crewlist"],
                         r["gcmid"], r["names_record"], r["names_title"],
                         r["name_sc"], r["email_sc"], r["phone_sc"], r["dept_sc"],
                         r["final"], r["confidence"], r["note"]])

    # Sheet 3: Summary by Project
    ws3 = out.create_sheet("Summary by Project")
    ws3.append(["Project", "Total rows", "Strong", "Good", "Weak", "Poor",
                "Check", "Needs Review", "% OK"])
    proj_stats = defaultdict(lambda: {"total": 0, "Strong": 0, "Good": 0,
                                       "Weak": 0, "Poor": 0, "Check": 0})
    for r in results:
        p = r["project"]
        proj_stats[p]["total"] += 1
        proj_stats[p][r["confidence"]] += 1

    for proj in sorted(proj_stats):
        s = proj_stats[proj]
        needs = s["Poor"] + s["Check"]
        ok_pct = round((s["Strong"] + s["Good"]) / s["total"] * 100, 1) if s["total"] else 0
        ws3.append([proj, s["total"], s["Strong"], s["Good"], s["Weak"],
                     s["Poor"], s["Check"], needs, ok_pct])

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    out.save(OUTPUT)

    # Summary counts
    counts = defaultdict(int)
    for r in results:
        counts[r["confidence"]] += 1

    needs_review = counts["Poor"] + counts["Check"]
    print(f"\nTotal rows scored:    {total_scored:,}")
    for label in ("Strong", "Good", "Weak", "Poor", "Check"):
        c = counts[label]
        pct = round(c / total_scored * 100, 1) if total_scored else 0
        print(f"  {label + ':':<22}{c:>6}  ({pct}%)")
    print(f"  {'Needs Review total:':<22}{needs_review:>6}")
    print(f"\nOutput: {OUTPUT}")


if __name__ == "__main__":
    main()
