import os
import pandas as pd
from glob import glob
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# === FOLDER SETUP ===
project_dir = os.path.dirname(os.path.abspath(__file__))
base_dir = os.path.join(project_dir, "Master_database")
sf_archive_dir = os.path.join(project_dir, "SF_Archive")

# === FILE PATHS ===
helper_path = os.path.join(base_dir, "Helper.xlsx")
map_input_path = os.path.join(base_dir, "new_titles_to_map.xlsx")
sflist_pattern = os.path.join(sf_archive_dir, "SFlist_*.xlsx")
output_path = os.path.join(base_dir, "new_titles_to_map.xlsx")

if not os.path.isdir(base_dir):
    raise FileNotFoundError(f"❌ Missing Master_database folder: {base_dir}")
if not os.path.isdir(sf_archive_dir):
    raise FileNotFoundError(f"❌ Missing SF_Archive folder: {sf_archive_dir}")

# === 1. UPDATE HELPER FROM FILLED-IN MAPPING ===
if os.path.exists(map_input_path):
    df_map = pd.read_excel(map_input_path, sheet_name="New Title-Project Pairs", engine="openpyxl")
    df_map["General Title"] = df_map["General Title"].astype(str).str.strip()
    df_map = df_map[~df_map["General Title"].isin(["", "nan", "NaN"])]

    if not df_map.empty:
        print(f"📥 Found {len(df_map)} completed mappings to update Helper.xlsx")

        # Load existing helper table
        df_helper = pd.read_excel(helper_path, sheet_name="Title conv", usecols="A:B", engine="openpyxl")
        existing_set = set(df_helper["Title-Project"].dropna().unique())

        # Filter only new rows
        df_to_add = df_map[~df_map["Title-Project"].isin(existing_set)][["Title-Project", "General Title"]]
        if not df_to_add.empty:
            df_new_helper = pd.concat([df_helper, df_to_add], ignore_index=True)
            with pd.ExcelWriter(helper_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df_new_helper.to_excel(writer, sheet_name="Title conv", index=False)
            print(f"✅ Added {len(df_to_add)} new rows to 'Title conv' in Helper.xlsx")
        else:
            print("ℹ️ No new mappings to add (already in Helper).")
    else:
        print("ℹ️ No completed mappings found in 'new_titles_to_map.xlsx'")
else:
    print("ℹ️ No previous 'new_titles_to_map.xlsx' found. Skipping update step.")

# === 2. GENERATE NEW MAPPINGS BASED ON LATEST SFLIST ===

# Find latest SFlist
sflist_files = glob(sflist_pattern)
if not sflist_files:
    print("❌ No SFlist_*.xlsx files found.")
    exit()
latest_sflist = max(sflist_files, key=os.path.getmtime)
print(f"📄 Using latest SFlist: {os.path.basename(latest_sflist)}")

# Extract current Title-Project pairs
df_sf = pd.read_excel(latest_sflist)
required_cols = {"Project", "Project job title", "Project department"}
if not required_cols.issubset(df_sf.columns):
    print(f"❌ Missing required columns: {required_cols - set(df_sf.columns)}")
    exit()

df_sf = df_sf[df_sf["Project job title"].notna()].copy()
df_sf["Project job title"] = df_sf["Project job title"].astype(str).str.strip()
df_sf["Project"] = df_sf["Project"].astype(str).str.strip()
df_sf["Title-Project"] = df_sf["Project job title"] + "--" + df_sf["Project"]


# Load current helper mappings
df_helper = pd.read_excel(helper_path, sheet_name="Title conv", usecols="A:B", engine="openpyxl")
mapped_set = set(df_helper["Title-Project"].dropna().unique())

# Identify unmapped title-projects and keep their departments
df_missing = df_sf[~df_sf["Title-Project"].isin(mapped_set)].drop_duplicates(subset=["Title-Project"])

# Build new table
df_new = df_missing[["Project department", "Project job title", "Project"]].copy()
df_new.columns = ["Project department", "Title", "Project"]
df_new["Title-Project"] = df_new["Title"] + "--" + df_new["Project"]
df_new["General Title"] = ""


# Load valid general titles
try:
    df_valid = pd.read_excel(helper_path, sheet_name="General Title", usecols="B", engine="openpyxl")
    df_valid = df_valid.rename(columns={df_valid.columns[0]: "General Title"}).dropna().drop_duplicates()
except Exception as e:
    print(f"❌ Could not read General Title tab: {e}")
    df_valid = pd.DataFrame(columns=["General Title"])

# === EXPORT WITHOUT FORMATTING ===
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df_new.to_excel(writer, sheet_name="New Title-Project Pairs", index=False)
    df_valid.to_excel(writer, sheet_name="Valid General Titles", index=False)

print(f"✅ Refreshed mapping file saved to: {output_path}")

# === ADD DROPDOWN VALIDATION FOR "General Title" ===
try:
    wb = load_workbook(output_path)
    ws_pairs = wb["New Title-Project Pairs"]
    ws_valid = wb["Valid General Titles"]

    # Determine data range in "Valid General Titles" column A (excluding header)
    last_row_valid = ws_valid.max_row
    if last_row_valid >= 2:
        formula = f"'Valid General Titles'!$A$2:$A${last_row_valid}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please select a value from the Valid General Titles list."
        dv.errorTitle = "Invalid General Title"
        dv.prompt = "Choose a General Title from the dropdown."
        dv.promptTitle = "General Title"
        ws_pairs.add_data_validation(dv)

        # Apply validation to column E ("General Title"), rows 2..max_row
        max_pairs_row = max(ws_pairs.max_row, 2)
        dv.add(f"E2:E{max_pairs_row}")

    wb.save(output_path)
    print("✅ Added dropdown validation to 'General Title' column.")
except Exception as e:
    print(f"⚠️ Could not add dropdown validation: {e}")
