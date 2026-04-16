"""New Combine + Preprocess pipeline.

Produces:
  New_Master_Database/CrewIndex.xlsx  — slim 26-col combined dataset + 6 derived sheets
  Updates New_Master_Database/CrewRegistry.xlsx — refreshes 6 auto-calculated columns
"""

import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────────

PROJECTS_JSON = Path("New_Master_Database/projects.json")
GCMID_MAP_FILE = Path("New_Master_Database/GCMID_Map.xlsx")
TITLEMAP_FILE = Path("New_Master_Database/TitleMap.xlsx")
REGISTRY_FILE = Path("New_Master_Database/CrewRegistry.xlsx")
HIST_DIR = Path("New_Master_Database/Historical")
SF_ARCHIVE = Path("SF_Archive")
OUTPUT_INDEX = Path("New_Master_Database/CrewIndex.xlsx")

# ── Constants ────────────────────────────────────────────────────────────────

BLOCKED_EMAILS = {"pioneer@crewcall.hu"}

NICKNAME_MAP = {
    "gabi": "gabriella", "zsuzsa": "zsuzsanna", "zsuzsi": "zsuzsanna", "gergo": "gergely",
    "kati": "katalin", "erzsi": "erzsebet", "bobe": "erzsebet", "bori": "borbala",
    "dani": "daniel", "moni": "monika", "zoli": "zoltan", "niki": "nikoletta",
    "pisti": "istvan", "magdi": "magdolna", "jr": "junior", "jrxx": "junior",
    "orsi": "orsolya", "ricsi": "richard", "gyuri": "gyorgy",
}

CREWINDEX_HEADERS = [
    "GCMID", "CM-Job", "Crew member id", "Crew list name", "Project",
    "Origin", "State", "Project department", "General Department",
    "Project job title", "General Title", "Department ID", "Title ID",
    "Deal type", "Business type",
    "Start date", "End date", "Project start date", "Project end date",
    "Daily fee", "Weekly fee", "Email", "Phone", "Surname", "Firstname",
    "Actual Name", "Actual Title", "Status",
]


# ── Utilities ────────────────────────────────────────────────────────────────

STATUS_VALUES_INLINE = '"Active,Retired,Foreign,External"'


def ensure_status_validation(ws):
    """Re-apply inline Status dropdown validation on CrewRegistry worksheet.

    Also fixes the Excel table column name if it still says "Retired",
    and clears the repairLoad flag that causes the repair dialog.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    header = [cell.value for cell in ws[1]]
    try:
        col_idx = header.index("Status") + 1
    except ValueError:
        return  # No Status column — nothing to do

    col_letter = get_column_letter(col_idx)

    # Fix table column name if it's still "Retired"
    if "CrewRegistry" in ws.tables:
        table = ws.tables["CrewRegistry"]
        for tc in table.tableColumns:
            if tc.name == "Retired":
                tc.name = "Status"
                break

    # Remove existing validations on the Status column
    to_remove = []
    for dv in ws.data_validations.dataValidation:
        for rng in dv.sqref.ranges:
            if str(rng).startswith(col_letter):
                to_remove.append(dv)
                break
    for dv in to_remove:
        ws.data_validations.dataValidation.remove(dv)

    # Apply fresh inline list validation
    dv = DataValidation(type="list", formula1=STATUS_VALUES_INLINE, allow_blank=True)
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
    ws.add_data_validation(dv)


def is_empty(val):
    return val is None or str(val).strip() in ("", "None", "nan")


def find_col(header, name):
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


def clean_int(val):
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return int(s)


def digits_only(val) -> str:
    if is_empty(val):
        return ""
    return re.sub(r"\D", "", str(val).strip())


def strip_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn")


def clean_token(text):
    if not isinstance(text, str):
        return ""
    text = strip_accents(text.lower())
    text = re.sub(r"[\"'''\"\"`,]", "", text)
    text = text.replace("-", " ")
    text = re.sub(r"[().]", "", text)
    return text.strip()


def tokenize_name(name):
    tokens = clean_token(name).split()
    result = set()
    for token in tokens:
        if len(token) >= 3 and token != "né":
            result.add(token)
            if token in NICKNAME_MAP:
                result.add(NICKNAME_MAP[token])
    return sorted(result)


def format_phone(phone_str):
    """Normalise to Hungarian 36-prefix format."""
    if not phone_str:
        return ""
    phone = re.sub(r"\D", "", str(phone_str))
    if len(phone) < 8:
        return ""
    if phone.startswith("36") or phone.startswith("00"):
        return phone
    if phone.startswith("06"):
        phone = "36" + phone[2:]
    elif phone.startswith("6"):
        phone = "36" + phone[1:]
    elif len(phone) == 9:
        phone = "36" + phone
    return phone


def get_val(row, col_idx):
    """Get value from row by column index, return None if missing."""
    if col_idx is None:
        return None
    val = row[col_idx]
    if is_empty(val):
        return None
    return val


def get_str(row, col_idx):
    val = get_val(row, col_idx)
    return str(val).strip() if val is not None else None


def get_float(row, col_idx):
    val = get_val(row, col_idx)
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ── Step 1: Load lookup tables ───────────────────────────────────────────────

def load_lookups():
    # GCMID map
    gcmid_map = {}
    wb = openpyxl.load_workbook(GCMID_MAP_FILE, read_only=True, data_only=True)
    ws = wb["GCMID_Map"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] is not None:
            gcmid_map[str(row[0]).strip()] = int(row[1])
    wb.close()

    # Title conv + General Title
    title_conv = {}
    title_to_dept = {}
    title_to_id = {}
    title_to_dept_id = {}
    wb = openpyxl.load_workbook(TITLEMAP_FILE, read_only=True, data_only=True)

    ws_tc = wb["Title conv"]
    for row in ws_tc.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            title_conv[str(row[0]).strip()] = str(row[1]).strip()

    ws_gt = wb["General Title"]
    for row in ws_gt.iter_rows(min_row=2, values_only=True):
        dept, title, dept_id, title_id = row[0], row[1], row[2], row[3]
        if title:
            t = str(title).strip()
            title_to_dept[t] = str(dept).strip() if dept else ""
            if title_id is not None:
                title_to_id[t] = int(title_id)
            if dept_id is not None:
                title_to_dept_id[t] = int(dept_id)
    wb.close()

    # Projects config
    project_dates = {}
    project_end_dates = {}
    pdata = json.loads(PROJECTS_JSON.read_text())
    for p in pdata["projects"]:
        name = p["name"]
        project_dates[name] = p.get("start_date") or ""
        project_end_dates[name] = p.get("end_date") or ""

    # CrewRegistry lookup by GCMID
    registry = {}
    wb = openpyxl.load_workbook(REGISTRY_FILE, read_only=True, data_only=True)
    ws = wb.active
    header = list(next(ws.iter_rows(max_row=1, values_only=True)))
    rc = {name: find_col(header, name) for name in [
        "CM ID", "Sure Name", "First Name", "Nick Name",
        "Actual Name", "Actual Title", "Status",
        "Actual Phone", "Actual Email",
        "Last Email", "Last Phone", "Last Department",
    ]}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cm_id = row[rc["CM ID"]]
        if is_empty(cm_id):
            continue
        cm_id_int = clean_int(cm_id)
        registry[cm_id_int] = {
            "sure_name": str(row[rc["Sure Name"]]).strip() if not is_empty(row[rc["Sure Name"]]) else None,
            "first_name": str(row[rc["First Name"]]).strip() if not is_empty(row[rc["First Name"]]) else None,
            "nick_name": str(row[rc["Nick Name"]]).strip() if not is_empty(row[rc["Nick Name"]]) else None,
            "actual_name": row[rc["Actual Name"]] if not is_empty(row[rc["Actual Name"]]) else None,
            "actual_title": row[rc["Actual Title"]] if not is_empty(row[rc["Actual Title"]]) else None,
            "status": str(row[rc["Status"]]).strip() if not is_empty(row[rc["Status"]]) else "Active",
            "actual_phone": row[rc["Actual Phone"]] if not is_empty(row[rc["Actual Phone"]]) else None,
            "actual_email": str(row[rc["Actual Email"]]).strip().lower() if not is_empty(row[rc["Actual Email"]]) else None,
            "last_email": str(row[rc["Last Email"]]).strip().lower() if not is_empty(row[rc["Last Email"]]) else None,
            "last_phone": row[rc["Last Phone"]] if not is_empty(row[rc["Last Phone"]]) else None,
            "last_department": str(row[rc["Last Department"]]).strip() if not is_empty(row[rc["Last Department"]]) else None,
        }
    wb.close()

    return gcmid_map, title_conv, title_to_dept, title_to_id, title_to_dept_id, project_dates, project_end_dates, registry


# ── Step 2: Read and normalise source rows ───────────────────────────────────

def read_source_file(path, origin, gcmid_map, title_conv, title_to_dept, title_to_id, title_to_dept_id, project_dates, project_end_dates):
    """Read one xlsx file and return normalised row dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return []

    header = all_rows[0]
    cols = {}
    for name in ["Crew member id", "Project", "Crew list name", "Surname", "Firstname",
                  "Email", "Crew email", "Mobile number", "Project job title",
                  "Project department", "State", "Deal type", "business_type",
                  "Daily fee", "Weekly fee", "Start date", "End date", "GCMID"]:
        cols[name] = find_col(header, name)

    rows_out = []
    for row in all_rows[1:]:
        # Filter draft/paused
        state_val = get_str(row, cols["State"])
        if state_val and state_val.lower() in ("draft", "paused"):
            continue

        crew_member_id = get_str(row, cols["Crew member id"])
        project = get_str(row, cols["Project"])
        crew_list_name = get_str(row, cols["Crew list name"])
        surname = get_str(row, cols["Surname"])
        firstname = get_str(row, cols["Firstname"])

        # Email: try Email first, fall back to Crew email
        email = get_str(row, cols["Email"])
        if not email:
            email = get_str(row, cols["Crew email"])
        if email:
            email = email.lower().strip()
        if email and email in BLOCKED_EMAILS:
            email = None

        phone_raw = get_str(row, cols["Mobile number"])
        phone_digits = digits_only(phone_raw)
        phone_int = int(phone_digits) if phone_digits and len(phone_digits) >= 8 else None

        project_job_title = get_str(row, cols["Project job title"])
        project_department = get_str(row, cols["Project department"])
        deal_type = get_str(row, cols["Deal type"])
        business_type = get_str(row, cols["business_type"])
        daily_fee = get_float(row, cols["Daily fee"])
        weekly_fee = get_float(row, cols["Weekly fee"])
        start_date = get_val(row, cols["Start date"])
        end_date = get_val(row, cols["End date"])

        # Derive CM-Job and GCMID
        cm_job = f"{crew_member_id}--{project}" if crew_member_id and project else None

        # Try GCMID from map first, then from file column
        gcmid = gcmid_map.get(cm_job) if cm_job else None
        if gcmid is None and cols.get("GCMID") is not None:
            gcmid_val = get_val(row, cols["GCMID"])
            if gcmid_val is not None:
                try:
                    gcmid = clean_int(gcmid_val)
                except (ValueError, TypeError):
                    pass

        # Title mapping
        title_key = f"{project_job_title}--{project}" if project_job_title and project else None
        general_title = title_conv.get(title_key) if title_key else None
        general_department = title_to_dept.get(general_title) if general_title else None

        proj_start = project_dates.get(project, "") if project else ""
        proj_end = project_end_dates.get(project, "") if project else ""

        rows_out.append({
            "gcmid": gcmid,
            "cm_job": cm_job,
            "crew_member_id": crew_member_id,
            "crew_list_name": crew_list_name,
            "project": project,
            "origin": origin,
            "state": state_val,
            "project_department": project_department,
            "general_department": general_department,
            "project_job_title": project_job_title,
            "general_title": general_title,
            "department_id": title_to_dept_id.get(general_title) if general_title else None,
            "title_id": title_to_id.get(general_title) if general_title else None,
            "deal_type": deal_type,
            "business_type": business_type,
            "start_date": start_date,
            "end_date": end_date,
            "project_start_date": proj_start,
            "project_end_date": proj_end,
            "daily_fee": daily_fee,
            "weekly_fee": weekly_fee,
            "email": email,
            "phone": phone_int,
            "surname": surname,
            "firstname": firstname,
            "source_file": path.name,
        })

    return rows_out


# ── Step 4: Derived sheets ───────────────────────────────────────────────────

def build_derived_sheets(all_rows, registry):
    """Build the 6 derived sheets from CrewIndex rows.

    All crew (regardless of Status) are included — Status is a view-level filter only.
    """

    by_gcmid = defaultdict(list)
    for r in all_rows:
        if r["gcmid"] is not None:
            by_gcmid[r["gcmid"]].append(r)

    # Tokenized Names
    token_rows = []
    for gcmid, rows in sorted(by_gcmid.items()):
        all_tokens = set()
        for r in rows:
            if r["crew_list_name"]:
                all_tokens.update(tokenize_name(r["crew_list_name"]))
        for tok in sorted(all_tokens):
            token_rows.append((gcmid, tok))

    # Phones
    phone_rows = []
    for gcmid, rows in sorted(by_gcmid.items()):
        seen = set()
        for r in rows:
            if r["phone"]:
                formatted = format_phone(str(r["phone"]))
                if formatted and formatted not in seen:
                    seen.add(formatted)
                    phone_rows.append((gcmid, formatted, r["phone"]))

    # Emails
    email_rows = []
    for gcmid, rows in sorted(by_gcmid.items()):
        seen = set()
        for r in rows:
            if r["email"] and r["email"] not in seen:
                seen.add(r["email"])
                email_rows.append((gcmid, r["email"]))

    # Actual Details (from CrewRegistry — all statuses)
    actual_rows = []
    for cm_id in sorted(registry.keys()):
        rec = registry[cm_id]
        actual_rows.append((
            cm_id,
            rec.get("actual_name"),
            rec.get("actual_title"),
            rec.get("actual_phone"),
            rec.get("actual_email"),
            rec.get("status", "Active"),
            rec.get("last_phone"),
            rec.get("last_email"),
        ))

    # Names
    names_rows = []
    for gcmid, rows in sorted(by_gcmid.items()):
        seen = set()
        for r in rows:
            if r["crew_list_name"] and r["crew_list_name"] not in seen:
                seen.add(r["crew_list_name"])
                names_rows.append((gcmid, r["crew_list_name"]))

    # General Departments
    dept_rows = []
    for gcmid, rows in sorted(by_gcmid.items()):
        seen = set()
        for r in rows:
            if r["general_department"] and r["general_department"] not in seen:
                seen.add(r["general_department"])
                dept_rows.append((gcmid, r["general_department"]))

    # ── Second pass: add registry-only members not in CrewIndex ──
    token_gcmids = {r[0] for r in token_rows}
    phone_gcmids = {r[0] for r in phone_rows}
    email_gcmids = {r[0] for r in email_rows}
    dept_gcmids = {r[0] for r in dept_rows}

    for cm_id in sorted(registry.keys()):
        rec = registry[cm_id]

        # Tokenized Names
        if cm_id not in token_gcmids:
            name_parts = " ".join(filter(None, [
                rec.get("first_name"), rec.get("sure_name"), rec.get("nick_name"),
            ]))
            if name_parts.strip():
                tokens = tokenize_name(name_parts)
                for tok in sorted(set(tokens)):
                    token_rows.append((cm_id, tok))

        # Phones
        if cm_id not in phone_gcmids:
            lp = rec.get("last_phone")
            if lp is not None:
                formatted = format_phone(str(lp))
                if formatted:
                    phone_rows.append((cm_id, formatted, str(lp)))

        # Emails
        if cm_id not in email_gcmids:
            le = rec.get("last_email")
            if le and le.strip().lower() not in BLOCKED_EMAILS:
                email_rows.append((cm_id, le.strip().lower()))

        # General Departments
        if cm_id not in dept_gcmids:
            ld = rec.get("last_department")
            if ld:
                dept_rows.append((cm_id, ld))

    return {
        "Tokenized Names": (["GCMID", "Token"], token_rows),
        "Phones": (["GCMID", "Phone", "Original Phone"], phone_rows),
        "Emails": (["GCMID", "Email"], email_rows),
        "Actual Details": (["CM ID", "Actual Name", "Actual Title", "Actual Phone", "Actual Email", "Status", "Last Phone", "Last Email"], actual_rows),
        "Names": (["GCMID", "Crew list name"], names_rows),
        "General Departments": (["GCMID", "General Department"], dept_rows),
    }


# ── Step 5: Update CrewRegistry auto columns ────────────────────────────────

def update_registry(all_rows, title_to_dept, title_to_id, project_dates):
    """Refresh the 6 auto columns in CrewRegistry.xlsx."""
    wb = openpyxl.load_workbook(REGISTRY_FILE)
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    col_cmid = header.index("CM ID") + 1
    col_lgt = header.index("Last General Title") + 1
    col_ldept = header.index("Last Department") + 1
    col_tflag = header.index("Title Flag") + 1
    col_lemail = header.index("Last Email") + 1
    col_lphone = header.index("Last Phone") + 1
    col_shows = header.index("Shows Worked") + 1
    col_atitle = header.index("Actual Title") + 1

    # Group rows by GCMID
    by_gcmid = defaultdict(list)
    for r in all_rows:
        if r["gcmid"] is not None:
            by_gcmid[r["gcmid"]].append(r)

    # Derive per GCMID
    derived = {}
    for gcmid, rows in by_gcmid.items():
        rows_sorted = sorted(rows, key=lambda r: r.get("project_start_date") or "", reverse=True)

        # Shows worked
        seen_proj = []
        seen_set = set()
        for r in rows_sorted:
            if r["project"] and r["project"] not in seen_set:
                seen_proj.append(r["project"])
                seen_set.add(r["project"])
        shows = " / ".join(seen_proj)

        # Last general title
        last_gt = ""
        for r in rows_sorted:
            if r["general_title"]:
                last_gt = r["general_title"]
                break

        last_dept = title_to_dept.get(last_gt, "") if last_gt else ""

        # Last email
        last_email = ""
        for r in rows_sorted:
            if r["email"]:
                last_email = r["email"]
                break

        # Last phone
        last_phone = None
        for r in rows_sorted:
            if r["phone"]:
                last_phone = r["phone"]
                break

        derived[gcmid] = {
            "last_gt": last_gt,
            "last_dept": last_dept,
            "last_email": last_email,
            "last_phone": last_phone,
            "shows": shows,
        }

    # Update rows
    stats = {"lgt": 0, "ldept": 0, "tflag": 0, "lemail": 0, "lphone": 0, "shows": 0,
             "promoted": 0, "lower_role": 0, "dept_change": 0}

    for row_idx in range(2, ws.max_row + 1):
        cm_id_val = ws.cell(row=row_idx, column=col_cmid).value
        if cm_id_val is None:
            continue
        cm_id = int(cm_id_val)
        d = derived.get(cm_id)
        if not d:
            continue

        # Last General Title — non-destructive: only overwrite if new value is non-empty
        old_lgt = ws.cell(row=row_idx, column=col_lgt).value or ""
        if d["last_gt"] and d["last_gt"] != str(old_lgt).strip():
            ws.cell(row=row_idx, column=col_lgt, value=d["last_gt"])
            stats["lgt"] += 1

        # Last Department — non-destructive
        old_ldept = ws.cell(row=row_idx, column=col_ldept).value or ""
        if d["last_dept"] and d["last_dept"] != str(old_ldept).strip():
            ws.cell(row=row_idx, column=col_ldept, value=d["last_dept"])
            stats["ldept"] += 1

        # Title Flag
        actual_title = ws.cell(row=row_idx, column=col_atitle).value
        actual_title_str = str(actual_title).strip() if actual_title else ""
        last_gt_str = d["last_gt"]
        title_flag = None

        if actual_title_str and last_gt_str:
            if actual_title_str == last_gt_str:
                title_flag = None
            else:
                at_dept = title_to_dept.get(actual_title_str, "")
                lg_dept = title_to_dept.get(last_gt_str, "")
                if at_dept and lg_dept and at_dept != lg_dept:
                    title_flag = "dept change"
                    stats["dept_change"] += 1
                elif at_dept == lg_dept and at_dept:
                    at_id = title_to_id.get(actual_title_str, 999)
                    lg_id = title_to_id.get(last_gt_str, 999)
                    if lg_id > at_id:
                        title_flag = "promoted"
                        stats["promoted"] += 1
                    elif lg_id < at_id:
                        title_flag = "lower role"
                        stats["lower_role"] += 1

        ws.cell(row=row_idx, column=col_tflag, value=title_flag)
        if title_flag:
            stats["tflag"] += 1

        # Last Email — non-destructive: only overwrite if new value is non-empty
        old_lemail = ws.cell(row=row_idx, column=col_lemail).value or ""
        if d["last_email"] and d["last_email"] != str(old_lemail).strip():
            ws.cell(row=row_idx, column=col_lemail, value=d["last_email"])
            stats["lemail"] += 1

        # Last Phone — non-destructive
        old_lphone = ws.cell(row=row_idx, column=col_lphone).value
        if d["last_phone"] is not None and d["last_phone"] != old_lphone:
            ws.cell(row=row_idx, column=col_lphone, value=d["last_phone"])
            stats["lphone"] += 1

        # Shows Worked — non-destructive
        old_shows = ws.cell(row=row_idx, column=col_shows).value or ""
        if d["shows"] and d["shows"] != str(old_shows).strip():
            ws.cell(row=row_idx, column=col_shows, value=d["shows"])
            stats["shows"] += 1

    ensure_status_validation(ws)

    try:
        wb.save(REGISTRY_FILE)
    except PermissionError:
        print("  ERROR: Could not save CrewRegistry.xlsx — please close the file and re-run.")
        raise

    return stats


# ── Step 6: Write CrewIndex.xlsx ─────────────────────────────────────────────

def write_crewindex(all_rows, derived_sheets, registry):
    """Write the CrewIndex.xlsx with main sheet + 6 derived sheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CrewIndex"

    # Header
    ws.append(CREWINDEX_HEADERS)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    for col in range(1, len(CREWINDEX_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"

    # Row data
    retired_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    foreign_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    external_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    status_fills = {"Retired": retired_fill, "Foreign": foreign_fill, "External": external_fill}

    for r in all_rows:
        reg = registry.get(r["gcmid"], {}) if r["gcmid"] else {}
        row_data = [
            r["gcmid"], r["cm_job"], r["crew_member_id"], r["crew_list_name"],
            r["project"], r["origin"], r["state"], r["project_department"],
            r["general_department"], r["project_job_title"], r["general_title"],
            r["department_id"], r["title_id"],
            r["deal_type"], r["business_type"], r["start_date"], r["end_date"],
            r["project_start_date"], r["project_end_date"],
            r["daily_fee"], r["weekly_fee"], r["email"], r["phone"],
            r["surname"], r["firstname"],
            reg.get("actual_name"), reg.get("actual_title"),
            reg.get("status", "Active"),
        ]
        ws.append(row_data)

    # Highlight rows by status
    num_cols = len(CREWINDEX_HEADERS)
    status_col = CREWINDEX_HEADERS.index("Status") + 1
    for row_idx in range(2, ws.max_row + 1):
        status_val = str(ws.cell(row=row_idx, column=status_col).value or "").strip()
        fill = status_fills.get(status_val)
        if fill:
            for col in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    # Column widths
    widths = {
        1: 8, 2: 28, 3: 16, 4: 20, 5: 18, 6: 10, 7: 12, 8: 20, 9: 20,
        10: 25, 11: 25, 12: 8, 13: 8, 14: 12, 15: 12, 16: 13, 17: 13,
        18: 13, 19: 13, 20: 12, 21: 12, 22: 28, 23: 15, 24: 20, 25: 20,
        26: 20, 27: 25, 28: 10,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Derived sheets
    for sheet_name, (headers, rows) in derived_sheets.items():
        ws_d = wb.create_sheet(title=sheet_name)
        ws_d.append(headers)
        for cell in ws_d[1]:
            cell.font = Font(bold=True)
        ws_d.freeze_panes = "A2"
        for row in rows:
            ws_d.append(list(row))

    try:
        wb.save(OUTPUT_INDEX)
    except PermissionError:
        print("  ERROR: Could not save CrewIndex.xlsx — please close the file and re-run.")
        raise


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("x_new_combine.py — CrewDB New Pipeline\n")

    # Step 1
    print("Loading lookup tables...")
    gcmid_map, title_conv, title_to_dept, title_to_id, title_to_dept_id, project_dates, project_end_dates, registry = load_lookups()
    print(f"  GCMID map: {len(gcmid_map)} entries")
    print(f"  Title conv: {len(title_conv)} entries")
    print(f"  Title→Dept: {len(title_to_dept)} entries")
    print(f"  CrewRegistry: {len(registry)} records")

    # Step 2
    print("\nReading source files...")
    pdata = json.loads(PROJECTS_JSON.read_text())
    all_rows = []
    hist_projects = 0
    hist_rows = 0
    sf_projects = 0
    sf_rows = 0
    filtered_count = 0
    warnings = []

    # Find latest SFlist
    sf_files = sorted(SF_ARCHIVE.glob("SFlist_*.xlsx"))
    sf_path = sf_files[-1] if sf_files else None
    if not sf_path:
        warnings.append("WARNING: No SFlist file found in SF_Archive/")

    # Pre-read SFlist once (if needed)
    sf_cache = None

    for p in pdata["projects"]:
        name = p["name"]
        state = p["state"]

        if state == "skip":
            continue

        if state == "historical":
            fname = "Historical_" + name.replace(" ", "_") + ".xlsx"
            fpath = HIST_DIR / fname
            if not fpath.exists():
                warnings.append(f"WARNING: Historical file not found for '{name}': {fname}")
                continue
            rows = read_source_file(fpath, "Historical", gcmid_map, title_conv, title_to_dept, title_to_id, title_to_dept_id, project_dates, project_end_dates)
            all_rows.extend(rows)
            hist_projects += 1
            hist_rows += len(rows)

        elif state == "live":
            if not sf_path:
                continue
            if sf_cache is None:
                sf_cache = read_source_file(sf_path, "SFlist", gcmid_map, title_conv, title_to_dept, title_to_id, title_to_dept_id, project_dates, project_end_dates)
            project_rows = [r for r in sf_cache if r["project"] == name]
            all_rows.extend(project_rows)
            if project_rows:
                sf_projects += 1
                sf_rows += len(project_rows)

    # Count filtered
    total_before_filter = hist_rows + sf_rows
    # (filtering already happened in read_source_file)

    # Sort
    all_rows.sort(key=lambda r: (r.get("project_start_date") or "", r.get("project") or "", r.get("crew_list_name") or ""), reverse=True)

    # Stats
    resolved = sum(1 for r in all_rows if r["gcmid"] is not None)
    unresolved = len(all_rows) - resolved
    mapped = sum(1 for r in all_rows if r["general_title"])
    unmapped = len(all_rows) - mapped

    for w in warnings:
        print(f"  {w}")

    # Step 4
    print("\nBuilding derived sheets...")
    derived = build_derived_sheets(all_rows, registry)

    # Step 5
    print("Updating CrewRegistry auto columns...")
    reg_stats = update_registry(all_rows, title_to_dept, title_to_id, project_dates)

    # Step 6
    print("Writing CrewIndex.xlsx...")
    write_crewindex(all_rows, derived, registry)

    # Step 7 — Build result and summary
    pct_r = round(resolved / len(all_rows) * 100, 1) if all_rows else 0
    pct_u = round(unresolved / len(all_rows) * 100, 1) if all_rows else 0
    pct_m = round(mapped / len(all_rows) * 100, 1) if all_rows else 0
    pct_um = round(unmapped / len(all_rows) * 100, 1) if all_rows else 0

    result = {
        "ok": True,
        "created_at": datetime.now().isoformat(),
        "meta": {
            "hist_projects": hist_projects,
            "hist_rows": hist_rows,
            "sf_projects": sf_projects,
            "sf_rows": sf_rows,
            "total_rows": len(all_rows),
            "gcmid_resolved": resolved,
            "gcmid_unresolved": unresolved,
            "title_mapped": mapped,
            "title_unmapped": unmapped,
            "registry_updated": reg_stats,
            "derived_sheets": {
                "tokenized_names": len(derived["Tokenized Names"][1]),
                "phones": len(derived["Phones"][1]),
                "emails": len(derived["Emails"][1]),
                "actual_details": len(derived["Actual Details"][1]),
                "names": len(derived["Names"][1]),
                "general_departments": len(derived["General Departments"][1]),
            },
            "warnings": warnings,
        },
        "output_files": [
            str(OUTPUT_INDEX),
            str(REGISTRY_FILE),
        ],
    }

    print(f"""
x_new_combine.py — CrewDB New Pipeline

Sources read:
  Historical files:  {hist_projects} projects, {hist_rows:,} rows
  SFlist (live):     {sf_projects} projects, {sf_rows:,} rows  (file: {sf_path.name if sf_path else 'N/A'})
  Total rows:        {len(all_rows):,}

GCMID resolution:
  Resolved:    {resolved:,} rows ({pct_r}%)
  Unresolved:  {unresolved:,} rows ({pct_u}%)

Title mapping:
  Mapped:      {mapped:,} rows ({pct_m}%)
  Unmapped:    {unmapped:,} rows ({pct_um}%)

CrewIndex.xlsx:
  Main sheet:          {len(all_rows):,} rows, {len(CREWINDEX_HEADERS)} cols
  Tokenized Names:     {len(derived['Tokenized Names'][1]):,} rows
  Phones:              {len(derived['Phones'][1]):,} rows
  Emails:              {len(derived['Emails'][1]):,} rows
  Actual Details:      {len(derived['Actual Details'][1]):,} rows
  Names:               {len(derived['Names'][1]):,} rows
  General Departments: {len(derived['General Departments'][1]):,} rows

CrewRegistry.xlsx auto columns refreshed:
  Last General Title updated: {reg_stats['lgt']}
  Last Department updated:    {reg_stats['ldept']}
  Title Flag set:             {reg_stats['tflag']} (promoted: {reg_stats['promoted']}, lower role: {reg_stats['lower_role']}, dept change: {reg_stats['dept_change']})
  Last Email updated:         {reg_stats['lemail']}
  Last Phone updated:         {reg_stats['lphone']}
  Shows Worked updated:       {reg_stats['shows']}

Output: {OUTPUT_INDEX}
        {REGISTRY_FILE} (auto columns refreshed)""")

    return result


if __name__ == "__main__":
    main()
